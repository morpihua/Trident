import csv
import random
import os
import io
import ast
import xlsxwriter
import openpyxl
import time
import math
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from math import sin, cos, atan2, sqrt
import copy
import zipfile
import xml.etree.ElementTree as ET
from xml.sax.saxutils import escape as xml_escape_util
import struct
import json
import base64
from datetime import datetime, timezone
import sys
import re
import ctypes
import uuid
from typing import Any, List, Dict, Optional, Tuple

# Список підтримуваних кольорів для експорту (RGB)
SUPPORTED_COLORS = {
    (77, 192, 77): ("ЗЕЛЕНИЙ", None),
    (0, 0, 0): ("ЧОРНИЙ", None),
    (229, 57, 114): ("МАЛИНОВИЙ", None),
    (255, 255, 255): ("БІЛИЙ", None),
    (54, 69, 77): ("СІРИЙ", None),
    (117, 79, 229): ("ФІОЛЕТОВИЙ", None),
    (255, 85, 51): ("ЧЕРВОНИЙ", None),
    (184, 82, 204): ("ФІОЛЕТОВИЙ2", None),
    (55, 82, 217): ("СИНІЙ", None),
    (89, 115, 128): ("СІРИЙ2", None),
    (179, 179, 179): ("СВІТЛОСІРИЙ", None),
    (37, 164, 254): ("БЛАКИТНИЙ", None),
    (255, 147, 39): ("ОРАНЖЕВИЙ", None),
    (176, 228, 103): ("САЛАТОВИЙ", None),
    (128, 99, 89): ("КОРИЧНЕВИЙ", None),
    (59, 213, 231): ("БІРЮЗОВИЙ", None),
    (35, 140, 131): ("БІРЮЗОВИЙ2", None),
    (255, 215, 13): ("ЖОВТИЙ", None),
    (242, 61, 61): ("ЧЕРВОНИЙ2", None),
    (244, 255, 129): ("ЛИМОННИЙ", None),
}

# SIDC для точок/орієнтирів
POINT_SIDC = "10016600006099000000"

# SIDC для ліній
LINES_SIDC_BY_COLOR = {
    "#f44336": "10062500001101010000",
    "#ffeb3b": "10012500001101020000",
    "#2196f3": "10032500001101020000",
    "#4caf50": "10042500001101010000",
    "#010101": "10066600001100000000",
}

# За замовчуванням для "невідомих" маршрутів
DEFAULT_LINE_COLOR_TUPLE = (255, 215, 13)
DEFAULT_LINE_HEX = '#ffeb3b'
DEFAULT_LINE_SIDC = LINES_SIDC_BY_COLOR[DEFAULT_LINE_HEX]


def get_line_sidc(hex_color):
    """SIDC для лінії за HEX"""
    hex_color = hex_color.lower()
    if hex_color in ["#ff5533", "#f44336", "#d32f2f", "#e53935", "#ff4c4c"]:
        return LINES_SIDC_BY_COLOR.get("#f44336", DEFAULT_LINE_SIDC)
    if hex_color in ["#ffd70d", "#ffeb3b"]:
        return LINES_SIDC_BY_COLOR.get("#ffeb3b", DEFAULT_LINE_SIDC)
    return LINES_SIDC_BY_COLOR.get(hex_color, DEFAULT_LINE_SIDC)


# Опціонально для кольорів у консолі
try:
    import colorama  # type: ignore

    colorama.init()
    COLORS_CONSOLE = {
        'W': colorama.Fore.YELLOW, 'E': colorama.Fore.RED,
        'P': '', 'D': colorama.Fore.CYAN, 'T': colorama.Fore.CYAN,
        'c': colorama.Fore.MAGENTA, 'o': colorama.Style.RESET_ALL,
    }
except ImportError:
    colorama = None
    COLORS_CONSOLE = {key: '' for key in ['W', 'E', 'P', 'D', 'T', 'c', 'o']}


def xml_escape(text_to_escape: Any) -> str:
    """Коректно екранує текст для вмісту XML."""
    if not isinstance(text_to_escape, str):
        text_to_escape = str(text_to_escape)
    return xml_escape_util(text_to_escape, entities={"'": "'", "\"": "\""})


class Tooltip:
    """Клас для створення підказок (tooltips) для віджетів Tkinter."""

    def __init__(self, widget, text, background="#313335", foreground="#EAEAEA"):
        self.widget = widget
        self.text = text
        self.background = background
        self.foreground = foreground
        self.tooltip_window = None
        widget.bind("<Enter>", self.enter)
        widget.bind("<Leave>", self.leave)

    def enter(self, event=None):
        x, y, _, _ = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + 25
        y += self.widget.winfo_rooty() + 20
        self.tooltip_window = tk.Toplevel(self.widget)
        self.tooltip_window.wm_overrideredirect(True)
        self.tooltip_window.wm_geometry(f"+{x}+{y}")
        label = tk.Label(self.tooltip_window, text=self.text, justify='left',
                         background=self.background, foreground=self.foreground, relief='solid', borderwidth=1,
                         font=("Courier New", 8, "normal"))
        label.pack(ipadx=2, ipady=2)

    def leave(self, event=None):
        if self.tooltip_window:
            self.tooltip_window.destroy()
        self.tooltip_window = None


class Base:
    """Базовий клас для логування з рівнями деталізації та опціональним виводом в GUI."""

    def __init__(self, verbosity: int = 0, gui_logger_func=None):
        self.verbosity = verbosity
        self.gui_logger_func = gui_logger_func

    def _log(self, level_char: str, message: str, *args):
        level_map_verbosity = {'E': -2, 'W': -1, 'P': 0, 'D': 1, 'T': 2}
        if self.verbosity < level_map_verbosity.get(level_char, 0):
            return

        prefix_map = {'E': 'ПОМИЛКА APQ: ', 'W': 'УВАГА APQ: ', 'P': 'APQ: ', 'D': 'НАЛАГОДЖЕННЯ APQ: ',
                      'T': 'TRACE APQ: '}
        log_message = prefix_map.get(level_char, '?') + (message % args if args else message)

        if self.gui_logger_func:
            is_error = level_char == 'E'
            is_warning = level_char == 'W'
            try:
                self.gui_logger_func(log_message, error=is_error, warning=is_warning)
            except TypeError:
                self.gui_logger_func(log_message)

        print(COLORS_CONSOLE.get(level_char, '') + log_message + COLORS_CONSOLE.get('o', ''), file=sys.stderr)

    def error(self, message: str, *args):
        self._log('E', message, *args)

    def warning(self, message: str, *args):
        self._log('W', message, *args)

    def print(self, message: str, *args):
        self._log('P', message, *args)

    def debug(self, message: str, *args):
        self._log('D', message, *args)

    def trace(self, message: str, *args):
        self._log('T', message, *args)

    def _load_raw(self, path_to_load: str) -> Optional[bytes]:
        if not os.path.isfile(path_to_load) or not os.access(path_to_load, os.R_OK):
            self.warning("Неможливо прочитати '%s'!", path_to_load)
            return None
        try:
            with open(path_to_load, 'rb') as f_in:
                raw_data_content = f_in.read()
            self.trace("Прочитано '%s': %d байт.", path_to_load, len(raw_data_content))
            return raw_data_content
        except IOError as e_io:
            self.warning("Помилка читання '%s': %s", path_to_load, e_io)
            return None


class ApqFile(Base):
    V100_HEADER_MAGIC_MASK = 0x50500000
    LDK_MAGIC_HEADER = 0x4C444B3A
    LDK_NODE_DATA_MAGIC = 0x00105555
    LDK_NODE_ADDITIONAL_DATA_MAGIC = 0x00205555
    LDK_NODE_MAGIC = 0x00015555
    LDK_NODE_LIST_MAGIC = 0x00025555
    LDK_NODE_TABLE_MAGIC = 0x00045555

    MAX_REASONABLE_STRING_LEN = 65536 * 2
    MAX_REASONABLE_ENTRIES = 100000

    def __init__(self, path=None, rawdata=None, file_type=None, rawname=None, rawts=None, verbosity=0, gui_logger_func=None):
        super().__init__(verbosity, gui_logger_func)
        self.path = path
        self.rawdata = rawdata
        self._file_type = file_type.lower() if file_type else None
        self.rawname = rawname
        self.rawts = rawts if rawts is not None else time.time()

        self.data_parsed = {}
        self.version = 0
        self.rawoffs = 0
        self.parse_successful = False
        load_success = False

        if self.path:
            file_name_local = os.path.basename(self.path)
            ext_match = file_name_local.lower().split('.')[-1] if '.' in file_name_local else ''
            aq_types = ["wpt", "set", "rte", "are", "trk", "ldk"]
            if ext_match in aq_types:
                self._file_type = ext_match
                self.rawdata = self._load_raw(self.path)
                if self.rawdata is None:
                    self.error(f"Не вдалося завантажити дані для {self.path}")
                    raise ValueError(f"Не вдалося завантажити {self.path}")
                try:
                    self.rawts = os.path.getmtime(self.path)
                except OSError:
                    self.rawts = time.time()
                load_success = True
            else:
                self.error("Невідомий тип файлу для шляху: %s!", self.path)
                raise ValueError(f"Unknown file type for {self.path}")
        elif self.rawdata is not None and self._file_type and self.rawname:
            valid_raw_types = ["wpt", "set", "rte", "are", "trk", "bin", "ldk"]
            if self._file_type not in valid_raw_types:
                self.error("Невідомий тип файлу: %s!", self._file_type)
                raise ValueError(f"Unknown raw type: {self._file_type}")
            self.path = self.rawname
            if self.rawts is None: self.rawts = time.time()
            load_success = True
        else:
            self.error("Неправильні параметри ApqFile!")
            raise ValueError("Illegal ApqFile params")

        if not load_success or self.rawdata is None:
            self.error("Дані не завантажено або відсутні для ApqFile.")
            return

        self.rawsize = len(self.rawdata)
        parser_method_name = f"_parse_{self._file_type}"

        if hasattr(self, parser_method_name) and callable(getattr(self, parser_method_name)):
            try:
                self.parse_successful = getattr(self, parser_method_name)()
            except Exception as e_parse:
                self.error(f"Виняток під час парсингу {self._file_type} ({self.path or self.rawname}): {e_parse}")
                self.parse_successful = False
        else:
            self.warning(f"Парсер для типу не знайдено: {self._file_type}")
            if self._file_type == "bin":
                self.data_parsed['raw_content_b64'] = base64.b64encode(self.rawdata).decode('ascii')
                self.parse_successful = True

    def _getval(self, val_type, arg=None):
        value = None
        type_map_struct = {
            'int': ('>i', 4), 'bool': ('>?', 1), 'byte': ('>b', 1), 'ubyte': ('>B', 1),
            'long': ('>q', 8), 'pointer': ('>Q', 8), 'double': ('>d', 8),
            'short': ('>h', 2), 'ushort': ('>H', 2)
        }
        if val_type in type_map_struct:
            struct_format, num_bytes = type_map_struct[val_type]
            if self.rawoffs + num_bytes > self.rawsize: return None
            try:
                value = struct.unpack(struct_format, self.rawdata[self.rawoffs: self.rawoffs + num_bytes])[0]
            except struct.error: return None
            self.rawoffs += num_bytes
            if val_type == 'bool': value = bool(value)
        elif val_type == 'int+raw':
            size_val = self._getval('int')
            if size_val is None or not (0 <= size_val <= self.MAX_REASONABLE_STRING_LEN * 10): return None
            if self.rawoffs + size_val > self.rawsize: return None
            value = base64.b64encode(self.rawdata[self.rawoffs: self.rawoffs + size_val]).decode('ascii')
            self.rawoffs += size_val
        elif val_type in ('raw', 'bin'):
            size = arg
            if size is None or not (0 <= size <= self.MAX_REASONABLE_STRING_LEN * 100): return None
            if self.rawoffs + size > self.rawsize: return None
            raw_bytes_read = self.rawdata[self.rawoffs: self.rawoffs + size]
            value = base64.b64encode(raw_bytes_read).decode('ascii') if val_type == 'raw' else raw_bytes_read
            self.rawoffs += size
        elif val_type == 'string':
            size = arg
            if size is None or not (0 <= size <= self.MAX_REASONABLE_STRING_LEN): return None
            if self.rawoffs + size > self.rawsize: return None
            try:
                value = self.rawdata[self.rawoffs: self.rawoffs + size].decode('utf-8')
            except UnicodeDecodeError:
                value = self.rawdata[self.rawoffs: self.rawoffs + size].decode('utf-8', errors='replace')
            self.rawoffs += size
        elif val_type == 'coords':
            int_val = self._getval('int')
            value = int_val * 1e-7 if int_val is not None else None
        elif val_type == 'height':
            int_val = self._getval('int')
            value = (None if int_val == -999999999 else int_val * 1e-3) if int_val is not None else None
        elif val_type == 'timestamp':
            long_val = self._getval('long')
            value = (None if long_val == 0 else long_val * 1e-3) if long_val is not None else None
        elif val_type == 'accuracy':
            int_val = self._getval('int')
            value = (None if int_val == 0 else int_val) if int_val is not None else None
        elif val_type == 'accuracy2':
            int_val = self._getval('int')
            value = (None if int_val == 0 else int_val * 1e-2) if int_val is not None else None
        elif val_type == 'pressure':
            int_val = self._getval('int')
            value = (None if int_val == 999999999 else int_val * 1e-3) if int_val is not None else None
        else: 
            return None
        return value
        
    def _getvalmulti(self, **kwargs_types):
        data_dict = {'_order': list(kwargs_types.keys())}
        for key, type_info in kwargs_types.items():
            arg = type_info[1] if isinstance(type_info, tuple) else None
            type_name = type_info[0] if isinstance(type_info, tuple) else type_info
            data_dict[key] = self._getval(type_name, arg)
        return data_dict

    def _check_header(self, *expected_file_versions):
        file_version = self._getval('int')
        if file_version is None: return None
        if (file_version & self.V100_HEADER_MAGIC_MASK) == self.V100_HEADER_MAGIC_MASK:
            file_version = (file_version & 0xff) + 100
        header_size = self._getval('int')
        if header_size is None or not (0 <= header_size <= min(self.rawsize, 1024)): return None
        self.version = file_version
        return header_size

    def _get_metadata(self):
        n_meta_entries = self._getval('int')
        if n_meta_entries is None or not (0 <= n_meta_entries <= self.MAX_REASONABLE_ENTRIES): return {}
        meta = {}
        for _ in range(n_meta_entries):
            name_len = self._getval('int')
            name_str = self._getval('string', name_len) if name_len and name_len > 0 else ("" if name_len == 0 else None)
            data_len_or_type = self._getval('int')
            data_value = None
            if data_len_or_type is not None:
                if data_len_or_type == -1: data_value = self._getval('bool')
                elif data_len_or_type == -2: data_value = self._getval('long')
                elif data_len_or_type == -3: data_value = self._getval('double')
                elif data_len_or_type == -4: data_value = self._getval('int+raw')
                elif data_len_or_type >= 0: data_value = self._getval('string', data_len_or_type)
            if name_str is not None: meta[name_str] = data_value
        return meta

    def _get_location(self):
        return self._getvalmulti(lon='coords', lat='coords', alt='height', ts='timestamp', acc='accuracy', bar='pressure')

    def _get_locations(self):
        n_loc = self._getval('int')
        if n_loc is None or not (0 <= n_loc <= self.MAX_REASONABLE_ENTRIES * 10): return []
        return [loc for _ in range(n_loc) if (loc := self._get_location()) is not None]

    def _get_waypoints(self):
        n_wp = self._getval('int')
        if n_wp is None or not (0 <= n_wp <= self.MAX_REASONABLE_ENTRIES): return []
        wp_list = []
        for _ in range(n_wp):
            meta, loc = self._get_metadata(), self._get_location()
            if meta is not None and loc is not None: wp_list.append({'meta': meta, 'location': loc})
        return wp_list

    def _get_segment(self):
        return {'meta': self._get_metadata(), 'locations': self._get_locations()}

    def _get_segments(self):
        n_seg = self._getval('int')
        if n_seg is None or not (0 <= n_seg <= self.MAX_REASONABLE_ENTRIES): return []
        return [seg for _ in range(n_seg) if (seg := self._get_segment()) is not None]

    def _parse_wpt(self):
        if self._check_header(2, 101) is None: return False
        self.data_parsed['meta'] = self._get_metadata()
        self.data_parsed['location'] = self._get_location()
        return all(k in self.data_parsed and self.data_parsed[k] is not None for k in ['meta', 'location'])

    def _parse_set(self):
        if self._check_header(2, 101) is None: return False
        self.data_parsed['meta'] = self._get_metadata()
        self.data_parsed['waypoints'] = self._get_waypoints()
        return all(k in self.data_parsed and self.data_parsed[k] is not None for k in ['meta', 'waypoints'])

    def _parse_rte(self): return self._parse_set()

    def _parse_are(self):
        if self._check_header(2) is None: return False
        self.data_parsed['meta'] = self._get_metadata()
        self.data_parsed['locations'] = self._get_locations()
        return all(k in self.data_parsed and self.data_parsed[k] is not None for k in ['meta', 'locations'])

    def _parse_trk(self):
        if self._check_header(2, 3, 101) is None: return False
        self.data_parsed['meta'] = self._get_metadata()
        self.data_parsed['waypoints'] = self._get_waypoints()
        self.data_parsed['segments'] = self._get_segments()
        return True

    def _parse_ldk(self):
        hdr = self._getvalmulti(magic='int', archVersion='int', rootOffset='pointer', res1='long', res2='long', res3='long', res4='long')
        print(f"DEBUG LDK header: {hdr}")
        if hdr.get('rootOffset') is None: return False
        self.data_parsed['root'] = self._get_node(hdr['rootOffset'])
        print(f"DEBUG LDK: Розібраний root: {self.data_parsed['root'] is not None}")
        return self.data_parsed.get('root') is not None

    def _get_node_data(self, initial_offset):
        self._seek(initial_offset)
        hdr = self._getvalmulti(magic='int', flags='int', totalSize='long', size='long', addOffset='pointer')
        if hdr.get('magic') != self.LDK_NODE_DATA_MAGIC: return None
        data_chunks = [main_data for main_data in [self._getval('bin', hdr.get('size'))] if main_data]
        offset = hdr.get('addOffset')
        while offset:
            self._seek(offset)
            add_hdr = self._getvalmulti(magic='int', size='long', addOffset='pointer')
            if add_hdr.get('magic') != self.LDK_NODE_ADDITIONAL_DATA_MAGIC: break
            if chunk := self._getval('bin', add_hdr.get('size')): data_chunks.append(chunk)
            offset = add_hdr.get('addOffset')
        return b"".join(data_chunks)

    def _get_node(self, offset, current_path="/", uid_for_path=None):
        if not (0 <= offset < self.rawsize): return None
        self._seek(offset)
        hdr = self._getvalmulti(magic='int', flags='int', metaOffset='pointer', res1='long')
        if hdr.get('magic') != self.LDK_NODE_MAGIC: return None

        self._seek(hdr['metaOffset'] + 0x20)
        node_meta = self._get_metadata()
        self._seek(offset + 24) # Return to position after header

        safe_name = re.sub(r'[\\/*?:"<>|]', '_', node_meta.get('name', '')) if node_meta else ''
        node_path = f"{current_path}{safe_name or f'UID{uid_for_path:08X}'}/"

        magic = self._getval('int')
        if magic == self.LDK_NODE_LIST_MAGIC:
            hdr2 = self._getvalmulti(nTotal='int', nChild='int', nData='int', addOffset='pointer')
            n_child, n_data = hdr2.get('nChild', 0), hdr2.get('nData', 0)
            n_empty = hdr2.get('nTotal', 0) - n_child - n_data
        elif magic == self.LDK_NODE_TABLE_MAGIC:
            hdr2 = self._getvalmulti(nChild='int', nData='int')
            n_child, n_data, n_empty = hdr2.get('nChild', 0), hdr2.get('nData', 0), 0
        else: return None

        child_defs = [d for i in range(n_child) if (d := self._getvalmulti(offset='pointer', uid='int', _ix=i))]
        if n_empty > 0: self._seek(self._tell() + n_empty * 12)
        data_defs = [d for i in range(n_data) if (d := self._getvalmulti(offset='pointer', uid='int', _ix=i))]

        files = []
        type_map = {0x65: 'wpt', 0x66: 'set', 0x67: 'rte', 0x68: 'trk', 0x69: 'are'}
        base_fn = os.path.splitext(os.path.basename(self.path or self.rawname))[0]
        for entry in data_defs:
            if file_bytes := self._get_node_data(entry['offset']):
                file_type_val, data = file_bytes[0], file_bytes[1:]
                if not data: continue
                type_str = type_map.get(file_type_val, 'bin')
                name = f"{base_fn}_{node_path.strip('/').replace('/', '_')}_UID{entry.get('uid', 0):08X}.{type_str}"
                files.append({'name': name, 'data_b64': base64.b64encode(data).decode('ascii'), 'type': type_str, 'size': len(data), 'order': entry['_ix']})

        nodes = [n for entry in child_defs if (n := self._get_node(entry['offset'], node_path, entry['uid']))]
        return {'path': node_path, 'nodes': nodes, 'files': files, 'meta': node_meta or {}}

    def _tell(self): return self.rawoffs
    def _seek(self, offset): self.rawoffs = offset
    def data(self): return self.get_parsed_data()

    def get_parsed_data(self):
        output = {'ts': self.rawts, 'type': self._file_type, 'path': self.path or self.rawname, 'file': os.path.basename(self.path or self.rawname or "unknown"), 'parse_successful': self.parse_successful}
        if self.parse_successful: output.update(self.data_parsed)
        return output

class Main:
    MAX_FILES: int = 100
    CSV_CHUNK_SIZE: int = 2000

    def __init__(self):
        self.program_version: str = "9.1_final"
        self.empty: str = "Не вибрано"
        self.file_ext: Optional[str] = None
        self.file_name: Optional[str] = None

        self.list_of_formats = [".geojson", ".kml", ".kmz", ".gpx", ".xlsx", ".csv", ".csv(макет)", ".scene"]
        self.supported_read_formats = [".kml", ".kmz", ".kme", ".gpx", ".xlsx", ".csv", ".scene", ".wpt", ".set", ".rte", ".are", ".trk", ".ldk", ".geojson"]

        self.numerations = ["За найближчими сусідами", "За змійкою", "За відстаню від кута", "За відстаню від границі", "За випадковістю"]
        self.translations = ["Не повертати", "На 90 градусів", "На 180 градусів", "На 270 градусів"]
        self.colors = {"Red": "#f44336", "Pink": "#e91e63", "Purple": "#9c27b0", "DeepPurple": "#673ab7", "Indigo": "#3f51b5", "Blue": "#2196f3", "Cyan": "#00bcd4", "Teal": "#009688", "Green": "#4caf50", "LightGreen": "#8bc34a", "Lime": "#cddc39", "Yellow": "#ffeb3b", "Amber": "#ffc107", "Orange": "#ff9800", "DeepOrange": "#ff5722", "Brown": "#795548", "BlueGrey": "#607d8b", "Black": "#010101", "White": "#ffffff"}
        self.color_options = ["Без змін"] + list(self.colors.keys())
        self.colors_en_ua = {"Red": "Червоний", "Pink": "Рожевий", "Purple": "Фіолетовий", "DeepPurple": "Темно-фіолетовий", "Indigo": "Індиго", "Blue": "Синій", "Cyan": "Блакитний", "Teal": "Бірюзовий", "Green": "Зелений", "LightGreen": "Салатовий", "Lime": "Лаймовий", "Yellow": "Жовтий", "Amber": "Бурштиновий", "Orange": "Помаранчевий", "DeepOrange": "Насичено-помаранчевий", "Brown": "Коричневий", "BlueGrey": "Синьо-сірий", "Black": "Чорний", "White": "Білий"}
        self.color_keyword_map = {"червоний": "Red", "рожевий": "Pink", "фіолетовий": "Purple", "синій": "Blue", "зелений": "Green", "жовтий": "Yellow", "оранжевий": "Orange", "коричневий": "Brown", "чорний": "Black", "білий": "White", "голубий": "Cyan"}
        self._palette_rgb = {name: (int(hx[1:3], 16), int(hx[3:5], 16), int(hx[5:7], 16)) for name, hx in self.colors.items()}
        
        self.file_list: List[Dict[str, Any]] = []
        self.list_is_visible: bool = False
        self.output_directory_path: str = self.empty
        self.font = ("Courier New", 11, "bold")
        self.C_BACKGROUND, self.C_SIDEBAR, self.C_BUTTON, self.C_BUTTON_HOVER, self.C_TEXT = "#2B2B2B", "#3C3C3C", "#556B2F", "#6B8E23", "#F5F5F5"
        self.C_ACCENT_SUCCESS, self.C_ACCENT_DONE, self.C_STATUS_DEFAULT, self.C_ACCENT_ERROR = "#6B8E23", "#FFBF00", "#4F4F4F", "#8B0000"

        self.main_window = tk.Tk()
        self.names_agree = tk.BooleanVar(value=False)
        self.exceptions_agree = tk.BooleanVar(value=False)
        self.chosen_numeration = tk.StringVar(value="За найближчими сусідами")
        self.chosen_translation = tk.StringVar(value="Не повертати")
        self.local_coords_agree = tk.BooleanVar(value=False)
        self.start_point = tk.StringVar(value="47.10000, 37.53000")
        self.start_point_numeric = [47.1, 37.53]

        self.main_window.title(f"Nexus Ultimate v{self.program_version}")
        self.main_window.configure(background=self.C_BACKGROUND)
        self.main_window.minsize(450, 120)
        self.main_window.geometry("450x120")
        self.main_window.protocol("WM_DELETE_WINDOW", self.exit)
        self.main_window.resizable(True, True)

        self._configure_styles()
        self._build_main_ui()
        
    def convert_color(self, color_value, mode='name'):
        """
        Перетворює колір у потрібний формат:
        - mode='hex': повертає hex (#rrggbb)
        - mode='name': повертає англомовну назву кольору із self.colors або 'White'
        """
        # Якщо вже hex у вигляді "#xxxxxx"
        if isinstance(color_value, str) and color_value.startswith("#") and len(color_value) == 7:
            if mode == 'hex':
                return color_value.lower()
            else:
                for k, v in self.colors.items():
                    if v.lower() == color_value.lower():
                        return k
                return 'White'
        # Якщо вже англійська або українська назва
        if isinstance(color_value, str):
            # Якщо укр, спробуй перекласти
            if hasattr(self, 'colors_en_ua') and color_value in self.colors_en_ua.values():
                for k, v in self.colors_en_ua.items():
                    if v == color_value:
                        color_value = k
                        break
            # Якщо англійська
            if color_value in self.colors:
                if mode == 'hex':
                    return self.colors[color_value]
                else:
                    return color_value
        # Якщо rgb-кортеж
        if isinstance(color_value, tuple) and len(color_value) == 3:
            hex_val = '#{:02x}{:02x}{:02x}'.format(*color_value)
            if mode == 'hex':
                return hex_val
            else:
                for k, v in self.colors.items():
                    if v.lower() == hex_val:
                        return k
                return 'White'
        return '#ffffff' if mode == 'hex' else 'White'

    def _configure_styles(self):
        style = ttk.Style(self.main_window)
        style.theme_use('clam')
        style.configure("TFrame", background=self.C_BACKGROUND)
        style.configure("Side.TFrame", background=self.C_SIDEBAR)
        style.configure("List.TFrame", background=self.C_SIDEBAR)
        style.configure('Icon.TButton', padding=5, borderwidth=0, relief='flat', background=self.C_BUTTON, foreground=self.C_TEXT, font=self.font)
        style.map('Icon.TButton', background=[('active', self.C_BUTTON_HOVER)], foreground=[('active', self.C_TEXT)])
        style.configure('Remove.TButton', background=self.C_SIDEBAR, foreground="#FF6347", font=("Courier New", 10, "bold"), relief='flat', borderwidth=0)
        style.map('Remove.TButton', background=[('active', "#4a4a4a")])
        style.configure("Toplevel", background=self.C_BACKGROUND)
        style.configure("TCheckbutton", background=self.C_BACKGROUND, foreground=self.C_TEXT, font=self.font, indicatorcolor=self.C_TEXT, selectcolor=self.C_BUTTON_HOVER)
        style.map("TCheckbutton", background=[('active', self.C_BACKGROUND)])
        style.configure("TLabel", background=self.C_BACKGROUND, foreground=self.C_TEXT, font=self.font)
        style.configure("List.TLabel", background=self.C_SIDEBAR, foreground=self.C_TEXT, font=("Courier New", 9))
        style.configure("Dark.TEntry", fieldbackground="#4F4F4F", foreground=self.C_TEXT, insertcolor=self.C_TEXT, bordercolor=self.C_SIDEBAR, font=("Courier New", 9))
        style.configure("TMenubutton", background="#4F4F4F", foreground=self.C_TEXT, font=("Courier New", 9), borderwidth=1, relief='raised', arrowcolor=self.C_TEXT)
        style.map("TMenubutton", background=[('active', "#646464")])

    def run(self):
        self.main_window.mainloop()

    def exit(self):
        if messagebox.askokcancel("Вихід", "Ви впевнені, що хочете вийти?"):
            self.main_window.destroy()

    def _build_main_ui(self):
        self.main_window.rowconfigure(1, weight=1)
        self.main_window.columnconfigure(0, weight=1)
        top_container = ttk.Frame(self.main_window)
        top_container.grid(row=0, column=0, sticky="ew", pady=(5, 0))
        top_container.columnconfigure(1, weight=1)
        
        left_sidebar = ttk.Frame(top_container, width=50, style="Side.TFrame")
        left_sidebar.grid(row=0, column=0, sticky="ns", padx=(5, 2))
        btn_lightbulb = ttk.Button(left_sidebar, text="i", style='Icon.TButton', command=self.show_info, width=2)
        btn_lightbulb.pack(pady=5, padx=5, fill='x')
        Tooltip(btn_lightbulb, "Про програму")
        btn_settings = ttk.Button(left_sidebar, text="S", style='Icon.TButton', command=self.open_settings, width=2)
        btn_settings.pack(pady=5, padx=5, fill='x')
        Tooltip(btn_settings, "Налаштування")

        center_frame = ttk.Frame(top_container)
        center_frame.grid(row=0, column=1, sticky="nsew")
        self.status_label = ttk.Label(center_frame, anchor="center", font=("Courier New", 14, "bold"), foreground=self.C_TEXT, relief='flat', padding=(0, 10))
        self.status_label.pack(fill="both", expand=True)
        self._update_status("ДОДАЙТЕ ФАЙЛИ", self.C_STATUS_DEFAULT)

        right_sidebar = ttk.Frame(top_container, width=50, style="Side.TFrame")
        right_sidebar.grid(row=0, column=2, sticky="ns", padx=(2, 5))
        self.btn_open_file = ttk.Button(right_sidebar, text="F", style='Icon.TButton', command=self.add_files_to_list, width=2)
        self.btn_open_file.pack(pady=5, padx=5, fill='x')
        Tooltip(self.btn_open_file, "Додати файли")
        self.play_button = ttk.Button(right_sidebar, text="▶", style='Icon.TButton', command=self.start_convertion, state="disabled", width=2)
        self.play_button.pack(pady=5, padx=5, fill='x')
        Tooltip(self.play_button, "Конвертувати все")

        self.list_container = ttk.Frame(self.main_window, style="List.TFrame")
        self.canvas = tk.Canvas(self.list_container, bg=self.C_SIDEBAR, highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.list_container, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas, style="List.TFrame")
        self.scrollable_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas_window = self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=scrollbar.set)
        self.canvas.pack(side="left", fill="both", expand=True, padx=(1, 0), pady=(1, 1))
        scrollbar.pack(side="right", fill="y", padx=(0, 1), pady=(1, 1))
        self.canvas.bind("<Configure>", self._on_canvas_configure)

    def _on_canvas_configure(self, event):
        self.canvas.itemconfig(self.canvas_window, width=event.width)

    def _redraw_file_list(self):
        for widget in self.scrollable_frame.winfo_children(): widget.destroy()
        for i, file_data in enumerate(self.file_list):
            item_frame = ttk.Frame(self.scrollable_frame, style="List.TFrame", padding=(5, 2))
            item_frame.pack(fill='x', expand=True)
            label_text = f"{i + 1}. {file_data['base_name']}"
            label = ttk.Label(item_frame, text=label_text[:40] + ('...' if len(label_text) > 40 else ''), style="List.TLabel", anchor='w')
            label.pack(side='left', fill='x', expand=True, padx=(0, 5))
            
            format_mb = ttk.Menubutton(item_frame, text=file_data['format_var'].get(), style="TMenubutton", width=10)
            menu = tk.Menu(format_mb, tearoff=0, bg=self.C_SIDEBAR, fg=self.C_TEXT, activebackground=self.C_BUTTON_HOVER)
            for fmt in self.list_of_formats:
                menu.add_radiobutton(label=fmt, variable=file_data['format_var'], value=fmt, command=lambda v=file_data['format_var'], b=format_mb, t=fmt: self._update_menubutton_text(v,b,t))
            format_mb['menu'] = menu
            format_mb.pack(side='left', padx=3)
            
            color_mb = ttk.Menubutton(item_frame, text=self.colors_en_ua.get(file_data['color_var'].get(), file_data['color_var'].get()), style="TMenubutton", width=12)
            menu = tk.Menu(color_mb, tearoff=0, bg=self.C_SIDEBAR, fg=self.C_TEXT, activebackground=self.C_BUTTON_HOVER)
            for color in self.color_options:
                menu.add_radiobutton(label=self.colors_en_ua.get(color, color), variable=file_data['color_var'], value=color, command=lambda v=file_data['color_var'], b=color_mb, t=color: self._update_menubutton_text(v,b,t))
            color_mb['menu'] = menu
            color_mb.pack(side='left', padx=3)
            
            remove_btn = ttk.Button(item_frame, text="X", style='Remove.TButton', width=2, command=lambda fd=file_data: self._remove_file(fd))
            remove_btn.pack(side='left', padx=(3, 0))
        
        status_text, color = (f"ГОТОВО: {len(self.file_list)} ФАЙЛ(ІВ)", self.C_ACCENT_SUCCESS) if self.file_list else ("ДОДАЙТЕ ФАЙЛИ", self.C_STATUS_DEFAULT)
        self._update_status(status_text, color)
        self.play_button.config(state="normal" if self.file_list else "disabled")
        
        if not self.file_list and self.list_is_visible:
            self.list_container.grid_forget()
            self.list_is_visible = False
            self.main_window.geometry("450x120")

    def _update_menubutton_text(self, var, button, value):
        var.set(value)
        button.config(text=self.colors_en_ua.get(value, value))

    def _remove_file(self, file_to_remove):
        self.file_list.remove(file_to_remove)
        self._redraw_file_list()

    def add_files_to_list(self):
        file_types = [("Підтримувані файли", " ".join(f"*{ext}" for ext in self.supported_read_formats)), ("Всі файли", "*.*")]
        paths = filedialog.askopenfilenames(filetypes=file_types, title="Виберіть файли для конвертації")
        if not paths: return
        
        for path in paths:
            if any(f['full_path'] == path for f in self.file_list):
                self._update_status(f"Файл вже у списку: {os.path.basename(path)}", warning=True)
                continue
            if len(self.file_list) >= self.MAX_FILES:
                messagebox.showwarning("Ліміт файлів", f"Максимальна кількість файлів ({self.MAX_FILES}) досягнута.")
                break
                
            default_format_map = {".kml": ".kml", ".kmz": ".kml", ".kme": ".kml", ".gpx": ".gpx", ".xlsx": ".xlsx", ".csv": ".csv", ".scene": ".geojson", ".geojson": ".geojson"}
            file_ext = os.path.splitext(path)[1].lower()
            default_export = default_format_map.get(file_ext, ".kml")
            
            file_data = {"full_path": path, "base_name": os.path.basename(path), "format_var": tk.StringVar(value=default_export), "color_var": tk.StringVar(value=self.color_options[0])}
            self.file_list.append(file_data)
        
        if self.file_list and not self.list_is_visible:
            self.list_container.grid(row=1, column=0, columnspan=3, sticky="nsew", padx=5, pady=(2, 5))
            self.main_window.rowconfigure(1, weight=3)
            self.list_is_visible = True
            self.main_window.geometry("650x400")
        
        self._redraw_file_list()

    def _update_status(self, text, color=None, error=False, warning=False):
        final_color = self.C_ACCENT_ERROR if error else (self.C_ACCENT_DONE if warning else (color if color else self.C_BACKGROUND))
        self.status_label.config(text=text.upper(), background=final_color)
        if self.main_window.winfo_exists(): self.main_window.update_idletasks()

    def _get_chunked_save_path(self, base_path, index):
        if index == 0: return base_path
        name, ext = os.path.splitext(base_path)
        return f"{name}({index + 1}){ext}"
        
    def show_info(self):
        messagebox.showinfo("Про програму", f"Nexus Ultimate v{self.program_version}\n\nПакетний конвертер геоданих.\n\nЧитання: {', '.join(self.supported_read_formats)}\nЗапис: {', '.join(self.list_of_formats)}")
        
    def open_settings(self):
        win = tk.Toplevel(self.main_window)
        win.title("Налаштування")
        win.configure(background=self.C_BACKGROUND)
        win.transient(self.main_window)
        win.grab_set()
        win.resizable(False, False)
        
        frame = ttk.Frame(win, padding=15)
        frame.pack(fill="both", expand=True)
        
        num_frame = ttk.LabelFrame(frame, text="Нумерація точок", padding=10)
        num_frame.pack(fill="x", expand=True, pady=5)
        ttk.Checkbutton(num_frame, text="Увімкнути нумерацію", variable=self.names_agree).pack(anchor="w", pady=(0, 5))
        ttk.Label(num_frame, text="Спосіб:").pack(anchor="w")
        ttk.Combobox(num_frame, textvariable=self.chosen_numeration, values=self.numerations, state="readonly").pack(fill="x", pady=(0, 5))
        ttk.Label(num_frame, text="Поворот:").pack(anchor="w")
        ttk.Combobox(num_frame, textvariable=self.chosen_translation, values=self.translations, state="readonly").pack(fill="x", pady=(0, 5))
        ttk.Checkbutton(num_frame, text="Виключити номери (30-40, 500-510)", variable=self.exceptions_agree).pack(anchor="w")

        loc_frame = ttk.LabelFrame(frame, text="Локальні координати (для CSV макету)", padding=10)
        loc_frame.pack(fill="x", expand=True, pady=10)
        entry = ttk.Entry(loc_frame, textvariable=self.start_point, justify="center", style="Dark.TEntry")
        def toggle_entry(): entry.pack(fill="x") if self.local_coords_agree.get() else entry.pack_forget()
        ttk.Checkbutton(loc_frame, text="Перевести в локальну площину (X,Y)", variable=self.local_coords_agree, command=toggle_entry).pack(anchor="w")
        toggle_entry()
        
        ttk.Button(frame, text="Закрити", command=win.destroy, style="Icon.TButton").pack(pady=(15, 0))
        win.update_idletasks()
        x = self.main_window.winfo_x() + (self.main_window.winfo_width() - win.winfo_width()) // 2
        y = self.main_window.winfo_y() + (self.main_window.winfo_height() - win.winfo_height()) // 2
        win.geometry(f"+{x}+{y}")
        
    def _process_data(self, content, color_override):
        if not content: return None
        processed = copy.deepcopy(content)
        if color_override != "Без змін":
            for item in processed: item['color'] = color_override
        if self.names_agree.get():
            points = [p for p in processed if p.get('geometry_type') == 'Point']
            others = [p for p in processed if p.get('geometry_type') != 'Point']
            if points: processed = self._apply_selected_numeration(points) + others
        return processed
        
    def start_convertion(self):
        if not self.file_list: return messagebox.showwarning("Увага", "Список файлів порожній.")
        if self.local_coords_agree.get():
            try: self.get_lat_lon_from_string(self.start_point.get())
            except ValueError as e: return messagebox.showerror("Помилка координат", str(e))
            
        readers = {ext: func for exts, func in [
            (('.kml', '.kme'), self.read_kml), (('.kmz',), self.read_kmz), (('.gpx',), self.read_gpx),
            (('.xlsx',), self.read_xlsx), (('.csv',), self.read_csv), (('.scene',), self.read_scene), 
            (('.geojson',), self.read_geojson), (('.wpt',), self.read_wpt), (('.set',), self.read_set), 
            (('.rte',), self.read_rte), (('.are',), self.read_are), (('.trk',), self.read_trk), (('.ldk',), self.read_ldk)
        ] for ext in exts}
        
        writers = {'.kml': self.create_kml, '.kmz': self.create_kmz, '.gpx': self.create_gpx, '.xlsx': self.create_xlsx, '.geojson': self.create_geojson, '.scene': self.create_scene}

        success_count = 0
        for i, file_data in enumerate(self.file_list):
            base_name, in_path = file_data['base_name'], file_data['full_path']
            self._update_status(f"ФАЙЛ {i + 1}/{len(self.file_list)}: {base_name}", self.C_BUTTON_HOVER)
            try:
                in_ext = os.path.splitext(in_path)[1].lower()
                content = readers.get(in_ext)(in_path) if readers.get(in_ext) else None
                if not content:
                    self._update_status(f"ПОМИЛКА ЧИТАННЯ: {base_name}", warning=True)
                    continue
                
                processed = self._process_data(content, file_data['color_var'].get())
                if not processed:
                    self._update_status(f"Помилка обробки: {base_name}", error=True)
                    continue

                out_format = file_data['format_var'].get().lower()
                clean_name = re.sub(r'\(\d+\)$', '', os.path.splitext(base_name)[0]).strip()
                save_path = filedialog.asksaveasfilename(
                    initialdir=self.output_directory_path if self.output_directory_path != self.empty else os.path.dirname(in_path),
                    initialfile=f"new_{clean_name}{out_format.replace('(макет)', '')}",
                    defaultextension=out_format.replace('(макет)', ''), title=f"Зберегти конвертований файл для {base_name}"
                )
                if not save_path:
                    self._update_status(f"СКАСОВАНО: {base_name}", warning=True)
                    continue
                
                self.output_directory_path = os.path.dirname(save_path)
                
                success = False
                if out_format in ('.csv', '.csv(макет)'):
                    is_layout = out_format == '.csv(макет)'
                    if self.create_csv(processed, save_path, layout_mode=is_layout): success_count += 1
                elif writers.get(out_format):
                    if writers[out_format](processed, save_path): success_count += 1
            except Exception as e:
                messagebox.showerror("Критична помилка", f"Не вдалося конвертувати {base_name}:\n{type(e).__name__}: {e}")
                self._update_status(f"КРИТИЧНА ПОМИЛКА: {base_name}", error=True)
                import traceback; traceback.print_exc()

        final_msg, final_color = (f"УСПІШНО ЗАВЕРШЕНО: {success_count}/{len(self.file_list)}", self.C_ACCENT_DONE) if success_count == len(self.file_list) else (f"ЗАВЕРШЕНО: {success_count}/{len(self.file_list)} УСПІШНО", self.C_ACCENT_DONE if success_count > 0 else self.C_ACCENT_ERROR)
        self._update_status(final_msg, final_color)
        if len(self.file_list) > 0:
            messagebox.showinfo("Завершено", f"Пакетна конвертація завершена.\nУспішно: {success_count} з {len(self.file_list)}.")

    def _read_specific_file(self, path, ext):
        try:
            apq = ApqFile(path=path, verbosity=2, gui_logger_func=self._update_status)
            if not apq.parse_successful:
                self._update_status(
                    f"Не вдалося розпарсити {ext.upper()}: {os.path.basename(path)}. "
                    f"Можливо, файл пошкоджений або має невірний формат.",
                    error=True
                )
                return None
            data = apq.data()
            # Для ApqFile .data() повертає dict, але якщо парсинг не вдався або даних немає — повертає порожній dict
            if not data or not data.get('parse_successful'):
                self._update_status(
                    f"Парсер не повернув даних для {ext.upper()} ({os.path.basename(path)}). "
                    f"Можливо, файл порожній або невірної структури.",
                    error=True
                )
                return None
            if apq._file_type == "bin":
                self._update_status(
                    f"Файл {os.path.basename(path)} є бінарним та не підтримується для конвертації.",
                    warning=True
                )
                return None
            return self._normalize_apq_content(data, path)
        except Exception as e:
            import traceback
            traceback.print_exc()
            self._update_status(f"Помилка {ext.upper()}: {os.path.basename(path)}: {e}", error=True)
            return None

    def read_wpt(self, path): return self._read_specific_file(path, ".wpt")
    def read_set(self, path): return self._read_specific_file(path, ".set")
    def read_rte(self, path): return self._read_specific_file(path, ".rte")
    def read_are(self, path): return self._read_specific_file(path, ".are")
    def read_trk(self, path): return self._read_specific_file(path, ".trk")

    def read_ldk(self, path):
        self._update_status(f"Читання LDK: {os.path.basename(path)}...", self.C_BUTTON_HOVER)
        all_normalized_content = []
        try:
            ldk_apq_file_instance = ApqFile(path=path, verbosity=2, gui_logger_func=self._update_status)
            if not ldk_apq_file_instance.parse_successful:
                self._update_status(
                    f"Не вдалося розпарсити LDK файл: {os.path.basename(path)}. "
                    f"Можливо, файл пошкоджений або має невірний формат.",
                    error=True
                )
                return None
            parsed_ldk_root_data = ldk_apq_file_instance.data()

            def extract_from_node(node_data):
                if not node_data:
                    print("DEBUG: contained_apq.data() for", ldk_file_entry.get('name'), ":", contained_apq.data())
                    return
                for ldk_file_entry in node_data.get('files', []):
                    print(f"DEBUG LDK: Вкладений файл: {ldk_file_entry.get('name')} тип: {ldk_file_entry.get('type')} розмір: {ldk_file_entry.get('size')}")
                    try:
                        inner_file_type = ldk_file_entry.get('type')
                        if inner_file_type == 'bin':
                            continue
                        file_bytes = base64.b64decode(ldk_file_entry['data_b64'])
                        contained_apq = ApqFile(rawdata=file_bytes, file_type=inner_file_type,
                                               rawname=ldk_file_entry.get('name'), verbosity=2,
                                               gui_logger_func=self._update_status)
                        if contained_apq.parse_successful:
                            normalized_data = self._normalize_apq_content(contained_apq.data(), path)
                            if normalized_data:
                                all_normalized_content.extend(normalized_data)
                        else:
                            self._update_status(
                                f"Вкладений файл LDK ({ldk_file_entry.get('name')}) не розпізнано.",
                                warning=True
                            )
                    except Exception as e:
                        import traceback
                        traceback.print_exc()
                        self._update_status(
                            f"Помилка розбору вкладеного файлу LDK ({ldk_file_entry.get('name')}): {type(e).__name__}: {e}",
                            warning=True
                        )
                        continue
                for child_node in node_data.get('nodes', []):
                    print(f"DEBUG LDK: Опрацьовуємо вузол: {node_data.get('path')} з файлами: {len(node_data.get('files', []))} і дочірніми: {len(node_data.get('nodes', []))}")
                    extract_from_node(child_node)

            if parsed_ldk_root_data and parsed_ldk_root_data.get('root'):
                extract_from_node(parsed_ldk_root_data['root'])
            else:
                self._update_status(
                    f"LDK файл {os.path.basename(path)} порожній або має невірну структуру.",
                    warning=True
                )
                return None
        except Exception as e:
            import traceback
            traceback.print_exc()
            self._update_status(
                f"Помилка читання LDK: {os.path.basename(path)}: {type(e).__name__}: {e}",
                error=True
            )
            return None
        return all_normalized_content if all_normalized_content else None

    def _normalize_apq_content(self, data, path):
        normalized_content = []
        print("DEBUG: type:", apq_type, "data keys:", list(data.keys()))
        print("DEBUG: waypoints:", data.get('waypoints'))
        if not data: return normalized_content
        apq_type = data.get('type')
        global_meta = data.get('meta', {})
        file_basename = os.path.basename(path)

        def _create_point(loc_data, meta_data, prefix, idx):
            if not loc_data or loc_data.get('lon') is None or loc_data.get('lat') is None: return None
            eff_meta = {**global_meta, **meta_data}
            name = eff_meta.get('name', f"{prefix}_{idx + 1}")
            color = self.convert_color(eff_meta.get('color', 'White'), 'name')
            desc = eff_meta.get('comment', eff_meta.get('description', ''))
            return {
                "name": name,
                "lat": loc_data['lat'],
                "lon": loc_data['lon'],
                "type": eff_meta.get('sym', 'Landmark'),
                "color": color,
                "description": desc,
                "geometry_type": "Point",
                "original_location_data": loc_data
            }

        if apq_type == 'wpt':
            if point := _create_point(data.get('location'), {}, 'Waypoint', 0):
                normalized_content.append(point)
        elif apq_type in ['set', 'rte']:
            waypoints = data.get('waypoints') or data.get('points') or data.get('locations') or []
            for i, wpt in enumerate(waypoints):
                if point := _create_point(wpt.get('location', wpt), wpt.get('meta', {}), global_meta.get('name', 'Wpt'), i):
                    normalized_content.append(point)

        elif apq_type == 'are':
            points = [loc for loc in data.get('locations', []) if loc and loc.get('lon') is not None]
            if len(points) >= 3:
                normalized_content.append({
                    'name': global_meta.get('name', 'Area'),
                    'geometry_type': 'Polygon',
                    'points_data': points,
                    'color': self.convert_color(global_meta.get('color', 'Blue'), 'name')
                })
        elif apq_type == 'trk':
            for i, poi in enumerate(data.get('waypoints', [])):
                if point := _create_point(poi.get('location'), poi.get('meta', {}), global_meta.get('name', 'POI'), i):
                    normalized_content.append(point)
            for i, seg in enumerate(data.get('segments', [])):
                points = [loc for loc in seg.get('locations', []) if loc and loc.get('lon') is not None]
                if len(points) >= 2:
                    seg_meta = {**global_meta, **seg.get('meta', {})}
                    normalized_content.append({
                        'name': seg_meta.get('name', f"{global_meta.get('name', 'Track')}_{i+1}"),
                        'geometry_type': 'LineString',
                        'points_data': points,
                        'color': self.convert_color(seg_meta.get('color', 'Red'), 'name')
                    })
        return normalized_content

    def _read_kml_from_content(self, kml_content, source_filename="KML"):
        result = []
        try:
            if isinstance(kml_content, bytes):
                if kml_content.startswith(b'\xef\xbb\xbf'): kml_content = kml_content[3:]
                kml_content = kml_content.decode('utf-8', 'replace')
            root = ET.fromstring(kml_content)
            ns = {'kml': re.match(r'\{([^}]+)\}', root.tag).group(1) if re.match(r'\{([^}]+)\}', root.tag) else 'http://www.opengis.net/kml/2.2'}
            for placemark in root.findall('.//kml:Placemark', ns):
                item_data = {}
                name_tag = placemark.find('kml:name', ns)
                item_data['name'] = name_tag.text.strip() if name_tag is not None and name_tag.text else 'KML Point'
                desc_tag = placemark.find('kml:description', ns)
                item_data['description'] = desc_tag.text.strip() if desc_tag is not None and desc_tag.text else ''
                # Color parsing logic
                style_url = placemark.findtext('kml:styleUrl', '', ns)
                style_node = placemark.find('kml:Style', ns)
                color_hex_str = None
                if style_url:
                    style_ref = root.find(f".//kml:Style[@id='{style_url.lstrip('#')}']", ns)
                    if style_ref is None:
                        style_map = root.find(f".//kml:StyleMap[@id='{style_url.lstrip('#')}']", ns)
                        if style_map is not None: style_ref = style_map.find(".//kml:key[text()='normal']/../kml:Style", ns)
                    if style_ref is not None: style_node = style_ref
                if style_node is not None:
                    for style_type in ['IconStyle', 'LineStyle', 'PolyStyle']:
                        color_tag = style_node.find(f'kml:{style_type}/kml:color', ns)
                        if color_tag is not None and color_tag.text:
                            c = color_tag.text.strip().lower()
                            if len(c) == 8: color_hex_str = f"#{c[6:8]}{c[4:6]}{c[2:4]}"
                            break
                item_data['color'] = self.convert_color(color_hex_str, 'name') if color_hex_str else 'White'
                # Geometry parsing
                point_coords = placemark.findtext('.//kml:Point/kml:coordinates', None, ns)
                line_coords = placemark.findtext('.//kml:LineString/kml:coordinates', None, ns)
                poly_coords = placemark.findtext('.//kml:Polygon/kml:outerBoundaryIs/kml:LinearRing/kml:coordinates', None, ns)
                if point_coords:
                    coords = point_coords.strip().split(',')
                    if len(coords) >= 2: item_data.update({'geometry_type': 'Point', 'lon': float(coords[0]), 'lat': float(coords[1])})
                elif line_coords:
                    points_data = [{'lon': float(p[0]), 'lat': float(p[1])} for p_str in line_coords.strip().split() if len(p := p_str.split(',')) >= 2]
                    if len(points_data) >= 2: item_data.update({'geometry_type': 'LineString', 'points_data': points_data})
                elif poly_coords:
                    points_data = [{'lon': float(p[0]), 'lat': float(p[1])} for p_str in poly_coords.strip().split() if len(p := p_str.split(',')) >= 2]
                    if len(points_data) >= 3: item_data.update({'geometry_type': 'Polygon', 'points_data': points_data})
                if 'geometry_type' in item_data: result.append(item_data)
        except Exception as e:
            self._update_status(f"Помилка KML: {source_filename}: {e}", error=True)
            return None
        return result

    def read_kml(self, path): return self._read_kml_from_content(open(path, 'rb').read(), os.path.basename(path))

    def read_kmz(self, path):
        try:
            with zipfile.ZipFile(path, 'r') as zf:
                kml_name = next((n for n in zf.namelist() if n.lower().endswith('.kml')), None)
                if kml_name: return self._read_kml_from_content(zf.read(kml_name), kml_name)
        except Exception as e:
            self._update_status(f"Помилка KMZ: {os.path.basename(path)}: {e}", error=True)
        return None
    
    def read_gpx(self, path):
        result = []
        try:
            tree = ET.parse(path)
            root = tree.getroot()
            ns = {'gpx': 'http://www.topografix.com/GPX/1/1'}
            for wpt in root.findall('gpx:wpt', ns):
                result.append({'name': wpt.findtext('gpx:name', 'GPX Point', ns), 'lon': float(wpt.get('lon')), 'lat': float(wpt.get('lat')), 'description': wpt.findtext('gpx:desc', '', ns), 'geometry_type': 'Point', 'color': 'White'})
            for trk in root.findall('gpx:trk', ns):
                for i, trkseg in enumerate(trk.findall('gpx:trkseg', ns)):
                    points_data = [{'lon': float(p.get('lon')), 'lat': float(p.get('lat'))} for p in trkseg.findall('gpx:trkpt', ns)]
                    if len(points_data) >= 2: result.append({'name': f"{trk.findtext('gpx:name', 'Track', ns)}_{i+1}", 'points_data': points_data, 'geometry_type': 'LineString', 'color': 'Red'})
        except Exception as e:
            self._update_status(f"Помилка GPX: {os.path.basename(path)}: {e}", error=True)
            return None
        return result

    def read_xlsx(self, path):
        result = []
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            for ws in wb.worksheets:
                if ws.max_row < 2: continue
                header = [str(cell.value).lower().strip() if cell.value else '' for cell in ws[1]]
                try:
                    lat_col = next(i for i, h in enumerate(header) if h in ['lat', 'latitude', 'широта', 'y'])
                    lon_col = next(i for i, h in enumerate(header) if h in ['lon', 'long', 'longitude', 'довгота', 'x'])
                    name_col = next((i for i, h in enumerate(header) if h in ['name', 'title', 'назва', 'id']), -1)
                except StopIteration: continue
                for row in ws.iter_rows(min_row=2, values_only=True):
                    try:
                        lat, lon = float(str(row[lat_col]).replace(',', '.')), float(str(row[lon_col]).replace(',', '.'))
                        if not (-90 <= lat <= 90 and -180 <= lon <= 180): continue
                        name = str(row[name_col]) if name_col != -1 and row[name_col] else f'{ws.title}_Point'
                        result.append({'name': name, 'lat': lat, 'lon': lon, 'geometry_type': 'Point', 'color': 'White'})
                    except (ValueError, TypeError, IndexError): continue
        except Exception as e:
            self._update_status(f"Помилка XLSX: {os.path.basename(path)}: {e}", error=True)
            return None
        return result

    def read_csv(self, path):
        result = []
        try:
            with open(path, 'r', encoding='utf-8-sig') as f:
                dialect = csv.Sniffer().sniff(f.read(4096))
                f.seek(0)
                reader = csv.DictReader(f, dialect=dialect)
                h_map = {h.lower().strip(): h for h in reader.fieldnames}
                lat_key = next((h_map[alias] for alias in ['lat', 'latitude', 'широта', 'y'] if alias in h_map), None)
                lon_key = next((h_map[alias] for alias in ['lon', 'long', 'longitude', 'довгота', 'x'] if alias in h_map), None)
                if not lat_key or not lon_key: return None
                name_key = next((h_map[alias] for alias in ['name', 'title', 'назва', 'id'] if alias in h_map), None)
                for row in reader:
                    try:
                        lat, lon = float(row[lat_key].replace(',', '.')), float(row[lon_key].replace(',', '.'))
                        name = row.get(name_key, 'CSV Point')
                        result.append({'name': name, 'lat': lat, 'lon': lon, 'geometry_type': 'Point', 'color': 'White'})
                    except (ValueError, TypeError, KeyError): continue
        except Exception as e:
            self._update_status(f"Помилка CSV: {os.path.basename(path)}: {e}", error=True)
            return None
        return result

    def read_scene(self, path):
        result = []
        try:
            with open(path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            for item in data.get("scene", {}).get("items", []):
                pos = item.get("position", {})
                if "lat" in pos and "lon" in pos:
                    result.append({"name": str(item.get("name", "SCENE Point")), "lon": float(pos["lon"]), "lat": float(pos["lat"]), "type": str(item.get("type", "Landmark")), "color": self.convert_color(str(item.get("color", "White")), "name"), "description": str(item.get("description", "")), "geometry_type": "Point"})
        except Exception as e:
            self._update_status(f"Помилка .scene: {os.path.basename(path)}: {e}", error=True)
            return None
        return result

    def read_geojson(self, path):
        result = []
        try:
            with open(path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            features = data.get("features", []) if data.get("type") == "FeatureCollection" else [data] if data.get("type") == "Feature" else []
            for feature in features:
                geom, props = feature.get("geometry", {}), feature.get("properties", {})
                if not geom or not props: continue
                item_base = {'name': props.get("name", "GeoJSON Feature"), 'color': self.convert_color(props.get("stroke", "#ffffff"), "name")}
                geom_type, coords = geom.get("type"), geom.get("coordinates")
                if geom_type == "Point" and coords and len(coords) >= 2:
                    item_base.update({'geometry_type': 'Point', 'lon': float(coords[0]), 'lat': float(coords[1])})
                    result.append(item_base)
                elif geom_type == "LineString" and coords and len(coords) >= 2:
                    item_base.update({'geometry_type': 'LineString', 'points_data': [{'lon': c[0], 'lat': c[1]} for c in coords]})
                    result.append(item_base)
                elif geom_type == "Polygon" and coords and len(coords[0]) >= 3:
                    item_base.update({'geometry_type': 'Polygon', 'points_data': [{'lon': c[0], 'lat': c[1]} for c in coords[0]]})
                    result.append(item_base)
        except Exception as e:
            self._update_status(f"Помилка GeoJSON: {os.path.basename(path)}: {e}", error=True)
            return None
        return result

    def _create_kml_string(self, contents, doc_name):
        kml = ET.Element("kml", xmlns="http://www.opengis.net/kml/2.2")
        doc = ET.SubElement(kml, "Document")
        ET.SubElement(doc, "name").text = doc_name
        style_map = {}
        for item in contents:
            color_name = item.get('color', 'White')
            if color_name not in style_map:
                hex_color = self.colors.get(color_name, "#ffffff").lstrip('#')
                kml_color = f"ff{hex_color[4:6]}{hex_color[2:4]}{hex_color[0:2]}"
                style = ET.SubElement(doc, "Style", id=f"style_{color_name}")
                style.append(ET.fromstring(f'<IconStyle><color>{kml_color}</color><Icon><href>http://maps.google.com/mapfiles/kml/pushpin/ylw-pushpin.png</href></Icon></IconStyle>'))
                style.append(ET.fromstring(f'<LineStyle><color>{kml_color}</color><width>2</width></LineStyle>'))
                style.append(ET.fromstring(f'<PolyStyle><color>7f{kml_color[2:]}</color></PolyStyle>'))
                style_map[color_name] = f"#style_{color_name}"
        for item in contents:
            pm = ET.SubElement(doc, "Placemark")
            ET.SubElement(pm, "name").text = xml_escape(item.get("name", "N/A"))
            if desc := item.get("description"): ET.SubElement(pm, "description").text = xml_escape(desc)
            ET.SubElement(pm, "styleUrl").text = style_map.get(item.get('color', 'White'))
            geom_type = item.get('geometry_type')
            if geom_type == "Point":
                coords = f"{item.get('lon', 0)},{item.get('lat', 0)},0"
                ET.SubElement(pm, "Point").append(ET.fromstring(f"<coordinates>{coords}</coordinates>"))
            elif geom_type in ["LineString", "Polygon"]:
                points_data = item.get('points_data', [])
                coords_str = " ".join(f"{p['lon']},{p['lat']},0" for p in points_data)
                if geom_type == "Polygon" and points_data and points_data[0] != points_data[-1]: coords_str += f" {points_data[0]['lon']},{points_data[0]['lat']},0"
                if geom_type == "LineString": geom = ET.SubElement(pm, "LineString")
                else: geom = ET.SubElement(ET.SubElement(ET.SubElement(pm, "Polygon"), "outerBoundaryIs"), "LinearRing")
                ET.SubElement(geom, "coordinates").text = coords_str
        ET.indent(kml, space="  ")
        return '<?xml version="1.0" encoding="UTF-8"?>\n' + ET.tostring(kml, encoding='unicode')

    def create_kml(self, contents, path):
        try:
            with open(path, "w", encoding="UTF-8") as f: f.write(self._create_kml_string(contents, os.path.splitext(os.path.basename(path))[0])); return True
        except IOError: return False
    
    def create_kmz(self, contents, path):
        try:
            with zipfile.ZipFile(path, 'w', zipfile.ZIP_DEFLATED) as zf: zf.writestr('doc.kml', self._create_kml_string(contents, "doc")); return True
        except: return False
    
    def create_gpx(self, contents, path):
        gpx = ET.Element('gpx', version="1.1", creator="Nexus", xmlns="http://www.topografix.com/GPX/1/1")
        for item in contents:
            if item.get('geometry_type') == 'Point':
                wpt = ET.SubElement(gpx, 'wpt', lat=str(item.get("lat")), lon=str(item.get("lon")))
                ET.SubElement(wpt, 'name').text = item.get("name")
            elif item.get('geometry_type') == 'LineString':
                trk, trkseg = ET.SubElement(gpx, 'trk'), ET.SubElement(ET.SubElement(gpx, 'trk'), 'trkseg')
                ET.SubElement(trk, 'name').text = item.get("name")
                for p in item.get('points_data', []): ET.SubElement(trkseg, 'trkpt', lat=str(p['lat']), lon=str(p['lon']))
        tree = ET.ElementTree(gpx)
        ET.indent(tree, space="  ")
        try:
            tree.write(path, encoding='utf-8', xml_declaration=True); return True
        except IOError: return False

    def create_xlsx(self, contents, path):
        try:
            with xlsxwriter.Workbook(path) as workbook:
                ws = workbook.add_worksheet("Data")
                headers = ["NAME", "LAT", "LON", "TYPE", "COLOR", "DESC", "GEOMETRY_TYPE", "WKT"]
                ws.write_row(0, 0, headers, workbook.add_format({'bold': True, 'bg_color': '#D9EAD3', 'border': 1}))
                for r, item in enumerate(contents, 1):
                    geom_type, wkt = item.get("geometry_type", ""), ""
                    if geom_type == "Point": wkt = f"POINT ({item.get('lon')} {item.get('lat')})"
                    elif geom_type in ["LineString", "Polygon"]:
                        pts = ", ".join(f"{p['lon']} {p['lat']}" for p in item.get('points_data', []))
                        if geom_type == "Polygon" and pts: wkt = f"POLYGON (({pts}, {item['points_data'][0]['lon']} {item['points_data'][0]['lat']}))"
                        else: wkt = f"LINESTRING ({pts})"
                    ws.write_row(r, 0, [item.get(k.lower(), '') for k in ['name', 'lat', 'lon', 'type', 'color', 'description']] + [geom_type, wkt])
                ws.autofit()
            return True
        except Exception: return False

    def create_geojson(self, contents, path):
        features = []
        for item in contents:
            geom_type = item.get('geometry_type')
            props = {k: v for k, v in item.items() if k not in ['points_data', 'geometry_type', 'original_location_data']}
            geometry = None
            if geom_type == 'Point':
                geometry = {"type": "Point", "coordinates": [item.get("lon", 0.0), item.get("lat", 0.0)]}
            elif geom_type in ['LineString', 'Polygon']:
                coords = [[p['lon'], p['lat']] for p in item.get('points_data', [])]
                if len(coords) >= (2 if geom_type == 'LineString' else 3):
                    if geom_type == 'Polygon' and coords[0] != coords[-1]: coords.append(coords[0])
                    geometry = {"type": geom_type, "coordinates": [coords] if geom_type == 'Polygon' else coords}
            if geometry: features.append({"type": "Feature", "properties": props, "geometry": geometry})
        if not features: return False
        try:
            with open(path, "w", encoding="UTF-8") as f: json.dump({"type": "FeatureCollection", "features": features}, f, indent=2, ensure_ascii=False); return True
        except IOError: return False

    def create_scene(self, contents_list, save_path):
        scene_items = []
        for item in contents_list:
            if item.get("geometry_type") == "Point":
                scene_items.append({
                    "color": item.get("color", "White"), "creationDate": int(time.time() * 1000), "description": item.get("description"),
                    "guid": str(uuid.uuid4()), "name": str(item.get("name", "N/A")),
                    "position": {"alt": 0.0, "lat": item.get("lat"), "lon": item.get("lon")}, # Lat/Lon swap for SCENE
                    "type": item.get("type", "Landmark")})
        scene_obj = {"scene": {"items": scene_items, "name": os.path.splitext(os.path.basename(save_path))[0]}, "version": 7}
        try:
            with open(save_path, "w", encoding="UTF-8") as f: json.dump(scene_obj, f, separators=(',', ':'), ensure_ascii=False); return True
        except IOError: return False

    def create_csv(self, contents, path, layout_mode=False):
        try:
            with open(path, "w", newline='', encoding="utf-8") as csvfile:
                if layout_mode:
                    writer = csv.writer(csvfile)
                    writer.writerow(["ID", "Name", "Lat", "Lon", "Alt", "X", "Y", "Z", "Tag"])
                    for i, item in enumerate(p for p in contents if p.get('geometry_type') == 'Point'):
                        lat, lon = item.get('lat', 0.0), item.get('lon', 0.0)
                        dx, dy = self.convert_to_local([lat, lon]) if self.local_coords_agree.get() else (0, 0)
                        writer.writerow([i, item.get('name', ''), lat, lon, 0, dx, dy, 0, "DefaultTag"])
                else:
                    writer = csv.writer(csvfile)
                    writer.writerow(["sidc", "name", "coordinates", "comment 1"])
                    for item in contents:
                        wkt, sidc, color_hex = "", POINT_SIDC, self.convert_color(item.get("color"), 'hex')
                        if item.get('geometry_type') == 'Point': wkt = f"POINT ({item.get('lon')} {item.get('lat')})"
                        elif item.get('geometry_type') in ['LineString', 'Polygon']:
                            sidc = get_line_sidc(color_hex)
                            pts = ", ".join(f"{p['lon']} {p['lat']}" for p in item.get('points_data', []))
                            if item.get('geometry_type') == 'Polygon' and pts: wkt = f"POLYGON (({pts}, {item['points_data'][0]['lon']} {item['points_data'][0]['lat']}))"
                            else: wkt = f"LINESTRING ({pts})"
                        if wkt: writer.writerow([sidc, item.get('name', ''), wkt, color_hex])
            return True
        except IOError: return False

    def calculate_distance(self, start, end):
        r = 6371
        lon1, lat1, lon2, lat2 = map(math.radians, [start[1], start[0], end[1], end[0]])
        a = sin((lat2 - lat1) / 2)**2 + cos(lat1) * cos(lat2) * sin((lon2 - lon1) / 2)**2
        return 2 * r * math.asin(sqrt(a)) * 1000

    def convert_to_local(self, point, start_point=None):
        start = start_point or self.start_point_numeric
        dx = self.calculate_distance(start, [start[0], point[1]]) * (1 if point[1] > start[1] else -1)
        dy = self.calculate_distance(start, [point[0], start[1]]) * (1 if point[0] > start[0] else -1)
        return [dx, dy]

    def get_lat_lon_from_string(self, s, sep=","):
        try: self.start_point_numeric = [round(float(p), 14) for p in s.replace(" ", "").split(sep)]
        except: raise ValueError("Некоректний формат координат")

    def _apply_selected_numeration(self, points):
        if not points: return []
        if len(points) == 1:
            points[0]['name'] = self.generate_free_numbers_list(1)[0]
            return points
        method = self.chosen_numeration.get()
        if method == "За найближчими сусідами": return self.apply_neighbor_numeration(points)
        if method == "За змійкою": return self.apply_snake_numeration(points)
        if method == "За відстаню від кута": return self.apply_two_axis_numeration(points)
        if method == "За відстаню від границі": return self.apply_one_axis_numeration(points)
        if method == "За випадковістю": return self.apply_random_numeration(points)
        return points

    def generate_free_numbers_list(self, count):
        nums, i = [], 1
        while len(nums) < count:
            if not (self.exceptions_agree.get() and (30 <= i <= 40 or 500 <= i <= 510)):
                nums.append(str(i))
            i += 1
        return nums

    def apply_random_numeration(self, points):
        nums = self.generate_free_numbers_list(len(points))
        random.shuffle(nums)
        for p, n in zip(points, nums): p['name'] = n
        return points

    def apply_neighbor_numeration(self, content):
        if not content: return []
        unvisited = copy.deepcopy(content)
        start_point = min(unvisited, key=lambda p: (p['lat'], p['lon']))
        ordered = [start_point]
        unvisited.remove(start_point)
        while unvisited:
            last = ordered[-1]
            next_point = min(unvisited, key=lambda p: self.calculate_distance((last['lat'], last['lon']), (p['lat'], p['lon'])))
            ordered.append(next_point)
            unvisited.remove(next_point)
        for i, p in enumerate(ordered): p['name'] = self.generate_free_numbers_list(len(ordered))[i]
        return ordered

    def apply_snake_numeration(self, content):
        if not content: return []
        points = copy.deepcopy(content)
        min_lat, max_lat = min(p['lat'] for p in points), max(p['lat'] for p in points)
        lat_range = max_lat - min_lat if max_lat != min_lat else 1e-6
        # Approximate number of rows, can be tuned
        num_rows = int(sqrt(len(points))) 
        points.sort(key=lambda p: (int((p['lat'] - min_lat) / lat_range * num_rows), p['lon'] if int((p['lat'] - min_lat) / lat_range * num_rows) % 2 == 0 else -p['lon']))
        for i, p in enumerate(points): p['name'] = self.generate_free_numbers_list(len(points))[i]
        return points

    def apply_one_axis_numeration(self, content):
        if not content: return []
        points = copy.deepcopy(content)
        trans = self.chosen_translation.get()
        if trans in ["На 90 градусів", "На 270 градусів"]:
            key, reverse = 'lat', trans == "На 270 градусів"
        else:
            key, reverse = 'lon', trans == "На 180 градусів"
        points.sort(key=lambda p: p[key], reverse=reverse)
        for i, p in enumerate(points): p['name'] = self.generate_free_numbers_list(len(points))[i]
        return points

    def apply_two_axis_numeration(self, content):
        if not content: return []
        points = copy.deepcopy(content)
        min_lat, max_lat = min(p['lat'] for p in points), max(p['lat'] for p in points)
        min_lon, max_lon = min(p['lon'] for p in points), max(p['lon'] for p in points)
        trans_map = {"Не повертати": (min_lat, min_lon), "На 90 градусів": (min_lat, max_lon), "На 180 градусів": (max_lat, max_lon), "На 270 градусів": (max_lat, min_lon)}
        corner = trans_map.get(self.chosen_translation.get(), (min_lat, min_lon))
        points.sort(key=lambda p: self.calculate_distance(corner, (p['lat'], p['lon'])))
        for i, p in enumerate(points): p['name'] = self.generate_free_numbers_list(len(points))[i]
        return points

if __name__ == "__main__":
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(1)
    except (ImportError, AttributeError, OSError):
        pass
    app = Main()
    app.run()
