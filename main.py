import csv
import random
import os
import io
import ast
import xlsxwriter  # type: ignore
import openpyxl  # type: ignore
import time
import math
import tkinter
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from math import sin, cos, atan2, sqrt
import copy
import zipfile
import xml.etree.ElementTree as ET
from xml.sax.saxutils import escape as xml_escape_util
import struct
import json
import base64
from datetime import datetime, timezone
import sys
import re
import ctypes
import uuid
from typing import Any, List, Dict, Optional, Tuple

# Список підтримуваних кольорів для експорту (RGB)
SUPPORTED_COLORS = {
    (77, 192, 77): ("ЗЕЛЕНИЙ", None),
    (0, 0, 0): ("ЧОРНИЙ", None),
    (229, 57, 114): ("МАЛИНОВИЙ", None),
    (255, 255, 255): ("БІЛИЙ", None),
    (54, 69, 77): ("СІРИЙ", None),
    (117, 79, 229): ("ФІОЛЕТОВИЙ", None),
    (255, 85, 51): ("ЧЕРВОНИЙ", None),
    (184, 82, 204): ("ФІОЛЕТОВИЙ2", None),
    (55, 82, 217): ("СИНІЙ", None),
    (89, 115, 128): ("СІРИЙ2", None),
    (179, 179, 179): ("СВІТЛОСІРИЙ", None),
    (37, 164, 254): ("БЛАКИТНИЙ", None),
    (255, 147, 39): ("ОРАНЖЕВИЙ", None),
    (176, 228, 103): ("САЛАТОВИЙ", None),
    (128, 99, 89): ("КОРИЧНЕВИЙ", None),
    (59, 213, 231): ("БІРЮЗОВИЙ", None),
    (35, 140, 131): ("БІРЮЗОВИЙ2", None),
    (255, 215, 13): ("ЖОВТИЙ", None),
    (242, 61, 61): ("ЧЕРВОНИЙ2", None),
    (244, 255, 129): ("ЛИМОННИЙ", None),
}

# SIDC для точок/орієнтирів
POINT_SIDC = "10016600006099000000"

# SIDC для ліній
LINES_SIDC_BY_COLOR = {
    "#f44336": "10062500001101010000",
    "#ffeb3b": "10012500001101020000",
    "#2196f3": "10032500001101020000",
    "#4caf50": "10042500001101010000",
    "#010101": "10066600001100000000",
}

# За замовчуванням для "невідомих" маршрутів
DEFAULT_LINE_COLOR_TUPLE = (255, 215, 13)
DEFAULT_LINE_HEX = '#ffeb3b'
DEFAULT_LINE_SIDC = LINES_SIDC_BY_COLOR[DEFAULT_LINE_HEX]


def get_line_sidc(hex_color):
    """SIDC для лінії за HEX"""
    hex_color = hex_color.lower()
    if hex_color in ["#ff5533", "#f44336", "#d32f2f", "#e53935", "#ff4c4c"]:
        return LINES_SIDC_BY_COLOR.get("#f44336", DEFAULT_LINE_SIDC)
    if hex_color in ["#ffd70d", "#ffeb3b"]:
        return LINES_SIDC_BY_COLOR.get("#ffeb3b", DEFAULT_LINE_SIDC)
    return LINES_SIDC_BY_COLOR.get(hex_color, DEFAULT_LINE_SIDC)


# Опціонально для кольорів у консолі
try:
    import colorama  # type: ignore

    colorama.init()
    COLORS_CONSOLE = {
        'W': colorama.Fore.YELLOW, 'E': colorama.Fore.RED,
        'P': '', 'D': colorama.Fore.CYAN, 'T': colorama.Fore.CYAN,
        'c': colorama.Fore.MAGENTA, 'o': colorama.Style.RESET_ALL,
    }
except ImportError:
    colorama = None
    COLORS_CONSOLE = {key: '' for key in ['W', 'E', 'P', 'D', 'T', 'c', 'o']}


def xml_escape(text_to_escape: Any) -> str:
    """Коректно екранує текст для вмісту XML."""
    if not isinstance(text_to_escape, str):
        text_to_escape = str(text_to_escape)
    return xml_escape_util(text_to_escape, entities={"'": "'", "\"": "\""})


class Tooltip:
    """Клас для створення підказок (tooltips) для віджетів Tkinter."""

    def __init__(self, widget, text, background="#313335", foreground="#EAEAEA"):
        self.widget = widget
        self.text = text
        self.background = background
        self.foreground = foreground
        self.tooltip_window = None
        widget.bind("<Enter>", self.enter)
        widget.bind("<Leave>", self.leave)

    def enter(self, event=None):
        x, y, _, _ = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + 25
        y += self.widget.winfo_rooty() + 20
        self.tooltip_window = tk.Toplevel(self.widget)
        self.tooltip_window.wm_overrideredirect(True)
        self.tooltip_window.wm_geometry(f"+{x}+{y}")
        label = tk.Label(self.tooltip_window, text=self.text, justify='left',
                         background=self.background, foreground=self.foreground, relief='solid', borderwidth=1,
                         font=("Courier New", 8, "normal"))
        label.pack(ipadx=2, ipady=2)

    def leave(self, event=None):
        if self.tooltip_window:
            self.tooltip_window.destroy()
        self.tooltip_window = None


class Base:
    """Базовий клас для логування з рівнями деталізації та опціональним виводом в GUI."""

    def __init__(self, verbosity: int = 0, gui_logger_func=None):
        self.verbosity = verbosity
        self.gui_logger_func = gui_logger_func

    def _log(self, level_char: str, message: str, *args):
        level_map_verbosity = {'E': -2, 'W': -1, 'P': 0, 'D': 1, 'T': 2}
        if self.verbosity < level_map_verbosity.get(level_char, 0):
            return

        prefix_map = {'E': 'ПОМИЛКА APQ: ', 'W': 'УВАГА APQ: ', 'P': 'APQ: ', 'D': 'НАЛАГОДЖЕННЯ APQ: ',
                      'T': 'TRACE APQ: '}
        log_message = prefix_map.get(level_char, '?') + (message % args if args else message)

        if self.gui_logger_func:
            is_error = level_char == 'E'
            is_warning = level_char == 'W'
            try:
                self.gui_logger_func(log_message, error=is_error, warning=is_warning)
            except TypeError:
                self.gui_logger_func(log_message)

        print(COLORS_CONSOLE.get(level_char, '') + log_message + COLORS_CONSOLE.get('o', ''), file=sys.stderr)

    def error(self, message: str, *args):
        self._log('E', message, *args)

    def warning(self, message: str, *args):
        self._log('W', message, *args)

    def print(self, message: str, *args):
        self._log('P', message, *args)

    def debug(self, message: str, *args):
        self._log('D', message, *args)

    def trace(self, message: str, *args):
        self._log('T', message, *args)

    def _load_raw(self, path_to_load: str) -> Optional[bytes]:
        if not os.path.isfile(path_to_load) or not os.access(path_to_load, os.R_OK):
            self.warning("Неможливо прочитати '%s'!", path_to_load)
            return None
        try:
            with open(path_to_load, 'rb') as f_in:
                raw_data_content = f_in.read()
            self.trace("Прочитано '%s': %d байт.", path_to_load, len(raw_data_content))
            return raw_data_content
        except IOError as e_io:
            self.warning("Помилка читання '%s': %s", path_to_load, e_io)
            return None


# --- Клас ApqFile ---
import os
import struct
import base64
import time
import re
from datetime import datetime, timezone

class ApqFile(Base):
    V100_HEADER_MAGIC_MASK = 0x50500000
    LDK_MAGIC_HEADER = 0x4C444B3A
    LDK_NODE_DATA_MAGIC = 0x00105555
    LDK_NODE_ADDITIONAL_DATA_MAGIC = 0x00205555
    LDK_NODE_MAGIC = 0x00015555
    LDK_NODE_LIST_MAGIC = 0x00025555
    LDK_NODE_TABLE_MAGIC = 0x00045555

    MAX_REASONABLE_STRING_LEN = 65536 * 2
    MAX_REASONABLE_ENTRIES = 100000

    def __init__(self, path=None, rawdata=None, file_type=None, rawname=None, rawts=None, verbosity=0,
                 gui_logger_func=None):
        super().__init__(verbosity, gui_logger_func)
        self.path = path
        self.rawdata = rawdata
        self._file_type = file_type.lower() if file_type else None
        self.rawname = rawname
        self.rawts = rawts if rawts is not None else time.time()

        self.data_parsed = {}
        self.version = 0
        self.rawoffs = 0
        self.parse_successful = False
        load_success = False

        # --- FILE LOADING ---
        if self.path:
            file_name_local = os.path.basename(self.path)
            ext_match = file_name_local.lower().split('.')[-1] if '.' in file_name_local else ''
            aq_types = ["wpt", "set", "rte", "are", "trk", "ldk"]
            if ext_match in aq_types:
                self._file_type = ext_match
                self.rawdata = self._load_raw(self.path)
                if self.rawdata is None:
                    self.error(f"Не вдалося завантажити дані для {self.path}")
                    raise ValueError(f"Не вдалося завантажити {self.path}")
                try:
                    self.rawts = os.path.getmtime(self.path)
                except OSError:
                    self.rawts = time.time()
                load_success = True
            else:
                self.error("Невідомий тип файлу для шляху: %s!", self.path)
                raise ValueError(f"Unknown file type for {self.path}")
        elif self.rawdata is not None and self._file_type and self.rawname:
            valid_raw_types = ["wpt", "set", "rte", "are", "trk", "bin", "ldk"]
            if self._file_type not in valid_raw_types:
                self.error("Невідомий тип файлу: %s!", self._file_type)
                raise ValueError(f"Unknown raw type: {self._file_type}")
            self.path = self.rawname
            if self.rawts is None: self.rawts = time.time()
            load_success = True
        else:
            self.error("Неправильні параметри ApqFile!")
            raise ValueError("Illegal ApqFile params")

        if not load_success or self.rawdata is None:
            self.error("Дані не завантажено або відсутні для ApqFile.")
            return

        self.rawsize = len(self.rawdata)
        parser_method_name = f"_parse_{self._file_type}"

        if hasattr(self, parser_method_name) and callable(getattr(self, parser_method_name)):
            try:
                self.parse_successful = getattr(self, parser_method_name)()
            except Exception as e_parse:
                self.error(f"Виняток під час парсингу {self._file_type} ({self.path or self.rawname}): {e_parse}")
                self.parse_successful = False
        else:
            self.warning(f"Парсер для типу не знайдено: {self._file_type}")
            if self._file_type == "bin":
                self.data_parsed['raw_content_b64'] = base64.b64encode(self.rawdata).decode('ascii')
                self.parse_successful = True

    def _getval(self, val_type, arg=None):
        original_offset = self.rawoffs
        value = None
        raw_bytes_read = b''
        type_map_struct = {
            'int': ('>i', 4), 'bool': ('>?', 1), 'byte': ('>b', 1), 'ubyte': ('>B', 1),
            'long': ('>q', 8), 'pointer': ('>Q', 8), 'double': ('>d', 8),
            'short': ('>h', 2), 'ushort': ('>H', 2)
        }
        if val_type in type_map_struct:
            struct_format, num_bytes = type_map_struct[val_type]
            if self.rawoffs + num_bytes > self.rawsize:
                return None
            raw_bytes_read = self.rawdata[self.rawoffs: self.rawoffs + num_bytes]
            try:
                value = struct.unpack(struct_format, raw_bytes_read)[0]
            except struct.error:
                return None
            self.rawoffs += num_bytes
            if val_type == 'bool': value = bool(value)
        elif val_type == 'int+raw':
            size_val = self._getval('int')
            if size_val is None or size_val < 0 or size_val > self.MAX_REASONABLE_STRING_LEN * 10:
                return None
            if self.rawoffs + size_val > self.rawsize:
                return None
            raw_bytes_read = self.rawdata[self.rawoffs: self.rawoffs + size_val]
            value = base64.b64encode(raw_bytes_read).decode('ascii')
            self.rawoffs += size_val
        elif val_type == 'raw' or val_type == 'bin':
            size = arg
            if size is None or size < 0 or size > self.MAX_REASONABLE_STRING_LEN * 100:
                return None
            if self.rawoffs + size > self.rawsize:
                return None
            raw_bytes_read = self.rawdata[self.rawoffs: self.rawoffs + size]
            value = base64.b64encode(raw_bytes_read).decode('ascii') if val_type == 'raw' else raw_bytes_read
            self.rawoffs += size
        elif val_type == 'string':
            size = arg
            if size is None or size < 0 or size > self.MAX_REASONABLE_STRING_LEN:
                return None
            if self.rawoffs + size > self.rawsize:
                return None
            raw_bytes_read = self.rawdata[self.rawoffs: self.rawoffs + size]
            try:
                value = raw_bytes_read.decode('utf-8')
            except UnicodeDecodeError:
                value = raw_bytes_read.decode('utf-8', errors='replace')
            self.rawoffs += size
        elif val_type == 'coords':
            int_val = self._getval('int')
            value = int_val * 1e-7 if int_val is not None else None
        elif val_type == 'height':
            int_val = self._getval('int')
            value = (None if int_val == -999999999 else int_val * 1e-3) if int_val is not None else None
        elif val_type == 'timestamp':
            long_val = self._getval('long')
            value = (None if long_val == 0 else long_val * 1e-3) if long_val is not None else None
        elif val_type == 'accuracy':
            int_val = self._getval('int')
            value = (None if int_val == 0 else int_val) if int_val is not None else None
        elif val_type == 'accuracy2':
            int_val = self._getval('int')
            value = (None if int_val == 0 else int_val * 1e-2) if int_val is not None else None
        elif val_type == 'pressure':
            int_val = self._getval('int')
            value = (None if int_val == 999999999 else int_val * 1e-3) if int_val is not None else None
        else:
            return None
        return value

    def _getvalmulti(self, **kwargs_types):
        data_dict = {'_order': list(kwargs_types.keys())}
        for key, type_info in kwargs_types.items():
            val = None
            arg_for_getval = None
            type_name_for_getval = type_info
            if isinstance(type_info, tuple):
                type_name_for_getval = type_info[0]
                if len(type_info) > 1:
                    arg_for_getval = type_info[1]
            val = self._getval(type_name_for_getval, arg_for_getval)
            data_dict[key] = val
        return data_dict

    def _check_header(self, *expected_file_versions):
        file_version = self._getval('int')
        if file_version is None:
            return None
        if (file_version & self.V100_HEADER_MAGIC_MASK) == self.V100_HEADER_MAGIC_MASK:
            file_version = (file_version & 0xff) + 100
        header_size = self._getval('int')
        if header_size is None or header_size < 0 or header_size > self.rawsize or header_size > 1024:
            return None
        self.version = file_version
        return header_size

    def _get_metadata(self):
        n_meta_entries = self._getval('int')
        if n_meta_entries is None or n_meta_entries < 0 or n_meta_entries > self.MAX_REASONABLE_ENTRIES:
            return {}
        meta = {}
        for _ in range(n_meta_entries):
            name_len = self._getval('int')
            name_str = self._getval('string', name_len) if name_len else None
            data_len_or_type = self._getval('int')
            data_value = None
            if data_len_or_type == -1:
                data_value = self._getval('bool')
            elif data_len_or_type == -2:
                data_value = self._getval('long')
            elif data_len_or_type == -3:
                data_value = self._getval('double')
            elif data_len_or_type == -4:
                data_value = self._getval('int+raw')
            elif data_len_or_type >= 0:
                data_value = self._getval('string', data_len_or_type)
            if name_str:
                meta[name_str] = data_value
        return meta

    def _get_location(self):
        loc = {}
        loc['lon'] = self._getval('coords')
        loc['lat'] = self._getval('coords')
        loc['alt'] = self._getval('height')
        loc['ts'] = self._getval('timestamp')
        loc['acc'] = self._getval('accuracy')
        loc['bar'] = self._getval('pressure')
        return loc

    def _get_waypoints(self):
        wp_list = []
        n_wp = self._getval('int')
        if n_wp is None or n_wp < 0 or n_wp > self.MAX_REASONABLE_ENTRIES:
            return []
        for _ in range(n_wp):
            meta = self._get_metadata()
            loc = self._get_location()
            if meta is None or loc is None:
                continue
            wp_list.append({'meta': meta, 'location': loc})
        return wp_list

    def _get_locations(self):
        loc_list = []
        n_loc = self._getval('int')
        if n_loc is None or n_loc < 0 or n_loc > self.MAX_REASONABLE_ENTRIES * 10:
            return []
        for _ in range(n_loc):
            loc = self._get_location()
            if loc is not None:
                loc_list.append(loc)
        return loc_list

    def _get_segment(self):
        seg_meta = self._get_metadata()
        n_loc = self._getval('int')
        locs_in_seg = []
        for _ in range(n_loc if n_loc else 0):
            loc = self._get_location()
            if loc is not None:
                locs_in_seg.append(loc)
        return {'meta': seg_meta, 'locations': locs_in_seg}

    def _get_segments(self):
        seg_list = []
        n_seg = self._getval('int')
        if n_seg is None or n_seg < 0 or n_seg > self.MAX_REASONABLE_ENTRIES:
            return []
        for _ in range(n_seg):
            seg_data = self._get_segment()
            if seg_data is not None:
                seg_list.append(seg_data)
        return seg_list

    def _parse_wpt(self):
        h_size = self._check_header(2, 101)
        if h_size is None:
            return False
        self.data_parsed['meta'] = self._get_metadata()
        self.data_parsed['location'] = self._get_location()
        return bool(self.data_parsed.get('meta') is not None and self.data_parsed.get('location') is not None)

    def _parse_set(self):
        h_size = self._check_header(2, 101)
        if h_size is None:
            return False
        self.data_parsed['meta'] = self._get_metadata()
        self.data_parsed['waypoints'] = self._get_waypoints()
        return bool(self.data_parsed.get('meta') is not None and self.data_parsed.get('waypoints') is not None)

    def _parse_rte(self):
        h_size = self._check_header(2, 101)
        if h_size is None:
            return False
        self.data_parsed['meta'] = self._get_metadata()
        self.data_parsed['waypoints'] = self._get_waypoints()
        return bool(self.data_parsed.get('meta') is not None and self.data_parsed.get('waypoints') is not None)

    def _parse_are(self):
        h_size = self._check_header(2)
        if h_size is None:
            return False
        self.data_parsed['meta'] = self._get_metadata()
        self.data_parsed['locations'] = self._get_locations()
        return bool(self.data_parsed.get('meta') is not None and self.data_parsed.get('locations') is not None)

    def _parse_trk(self):
        h_size = self._check_header(2, 3, 101)
        if h_size is None:
            return False
        self.data_parsed['meta'] = self._get_metadata()
        self.data_parsed['waypoints'] = self._get_waypoints()
        self.data_parsed['segments'] = self._get_segments()
        return True

    def _parse_ldk(self):
        hdr = self._getvalmulti(magic='int', archVersion='int', rootOffset='pointer',
                                res1='long', res2='long', res3='long', res4='long')
        if None in [hdr.get('magic'), hdr.get('archVersion'), hdr.get('rootOffset')]:
            return False
        root_offset_val = hdr.get('rootOffset')
        self.data_parsed['root'] = self._get_node(root_offset_val)
        return self.data_parsed.get('root') is not None

    def _get_node_data(self, initial_offset):
        self._seek(initial_offset)
        hdr = self._getvalmulti(magic='int', flags='int', totalSize='long', size='long', addOffset='pointer')
        if hdr.get('magic') != self.LDK_NODE_DATA_MAGIC:
            return None
        data_chunks = []
        main_data_block = self._getval('bin', hdr.get('size'))
        if main_data_block is not None:
            data_chunks.append(main_data_block)
        current_add_offset_val = hdr.get('addOffset')
        while current_add_offset_val:
            self._seek(current_add_offset_val)
            add_hdr = self._getvalmulti(magic='int', size='long', addOffset='pointer')
            if add_hdr.get('magic') != self.LDK_NODE_ADDITIONAL_DATA_MAGIC:
                break
            additional_data_block = self._getval('bin', add_hdr.get('size'))
            if additional_data_block is not None:
                data_chunks.append(additional_data_block)
            current_add_offset_val = add_hdr.get('addOffset')
        return b"".join(data_chunks)

    def _get_node(self, offset, current_path_prefix="/", uid_for_path=None):
        # Переміщаємось у файл
        if offset >= self.rawsize:
            self.error(f"Некоректний offset вузла LDK: {offset}")
            return None

        self._seek(offset)
        hdr = self._getvalmulti(magic='int', flags='int', metaOffset='pointer', res1='long')
        if None in [hdr.get('magic'), hdr.get('metaOffset')]:
            self.error("Не вдалося прочитати заголовок вузла LDK (обов'язкові поля).")
            return None
        if hdr.get('magic') != self.LDK_NODE_MAGIC:
            self.warning('Невідомий LDK node magic 0x%08x.', hdr.get('magic'))
            return None

        meta_offset_val = hdr.get('metaOffset')
        prev_offs = self._tell()
        self._seek(meta_offset_val + 0x20)
        node_meta = self._get_metadata()
        self._seek(prev_offs)

        node_path = current_path_prefix
        if uid_for_path is not None:
            node_name_from_meta = node_meta.get('name') if node_meta else None
            safe_node_name = re.sub(r'[\\/*?:"<>|]', '_', node_name_from_meta) if node_name_from_meta else None
            node_path += f"{safe_node_name}/" if safe_node_name else f"UID{uid_for_path:08X}/"

        node_entries_magic = self._getval('int')
        if node_entries_magic is None:
            self.error("Не вдалося прочитати magic для записів вузла LDK.")
            return None

        self.debug('LDK node path=%s, nodeEntriesMagic=0x%08x', node_path, node_entries_magic)
        node_obj = {'path': node_path, 'nodes': [], 'files': [], 'meta': node_meta if node_meta else {}}
        n_child, n_data, n_empty = 0, 0, 0

        if node_entries_magic == self.LDK_NODE_LIST_MAGIC:
            list_hdr = self._getvalmulti(nTotal='int', nChild='int', nData='int', addOffset='pointer')
            if None in [list_hdr.get('nTotal'), list_hdr.get('nChild'), list_hdr.get('nData')]:
                return None
            n_child, n_data = list_hdr.get('nChild', 0), list_hdr.get('nData', 0)
            n_empty = list_hdr.get('nTotal', 0) - n_child - n_data
        elif node_entries_magic == self.LDK_NODE_TABLE_MAGIC:
            self.warning("LDK: Обробка вузла типу 'таблиця' (0x00045555) може бути неповною.")
            table_hdr_simple = self._getvalmulti(nChild='int', nData='int')
            if table_hdr_simple.get('nChild') is not None and table_hdr_simple.get('nData') is not None:
                n_child, n_data = table_hdr_simple.get('nChild', 0), table_hdr_simple.get('nData', 0)
            else:
                self.error("LDK: Не вдалося прочитати nChild/nData для вузла-таблиці. Структура невідома.")
                return None
        else:
            self.warning('Неправильний LDK node entries magic 0x%08x.', node_entries_magic)
            return None

        entry_size = 12
        child_defs, data_defs = [], []
        for i in range(n_child):
            d = self._getvalmulti(offset='pointer', uid='int')
            if None in [d.get('offset'), d.get('uid')]:
                self.error(f"Помилка читання child_def {i}")
                return None
            d['_ix'] = i
            child_defs.append(d)
            self.trace('LDK childDef[%d]: off=0x%x uid=0x%x', i, d['offset'], d['uid'])

        if n_empty < 0:
            self.warning(f"Негативна кількість порожніх записів ({n_empty}) у вузлі LDK.")
            n_empty = 0
        bytes_to_skip = n_empty * entry_size
        if self._tell() + bytes_to_skip > self.rawsize:
            self.error(f"LDK: Спроба пропустити порожні записи виходить за межі файлу.")
            return None
        self._seek(self._tell() + bytes_to_skip)

        for i in range(n_data):
            d = self._getvalmulti(offset='pointer', uid='int')
            if None in [d.get('offset'), d.get('uid')]:
                self.error(f"Помилка читання data_def {i}")
                return None
            d['_ix'] = i
            data_defs.append(d)
            self.trace('LDK dataDef[%d]: off=0x%x uid=0x%x', i, d['offset'], d['uid'])

        # Розбір дочірніх вузлів
        for entry_def in sorted(child_defs, key=lambda x: x['_ix']):
            if entry_def['offset'] == 0:
                self.warning(f"LDK: Нульовий offset для дочірнього вузла UID {entry_def['uid']}. Пропускається.")
                continue
            child_node = self._get_node(entry_def['offset'], node_path, entry_def['uid'])
            if child_node is None:
                self.error(f"Помилка парсингу дочірнього вузла LDK (offset {entry_def['offset']}).")
                continue
            child_node['order'] = entry_def['_ix']
            node_obj['nodes'].append(child_node)

        # Розбір файлів (даних) цього вузла
        type_map_ldk = {0x65: 'wpt', 0x66: 'set', 0x67: 'rte', 0x68: 'trk', 0x69: 'are'}
        ldk_original_filename = self.path or self.rawname or "unknown.ldk"
        ldk_base_fn_for_contained = os.path.splitext(os.path.basename(ldk_original_filename))[0]

        for entry_def in sorted(data_defs, key=lambda x: x['_ix']):
            if entry_def['offset'] == 0:
                self.warning(f"LDK: Нульовий offset для файлу UID {entry_def['uid']}. Пропускається.")
                continue
            file_bytes = self._get_node_data(entry_def['offset'])
            if file_bytes is None or not file_bytes:
                self.warning(f"Пропущено порожній/пошкоджений файл у LDK (UID {entry_def.get('uid', 'N/A')})")
                continue
            file_type_val = file_bytes[0]
            actual_data_bytes = file_bytes[1:]
            if not actual_data_bytes:
                self.warning(f"LDK: Файл UID {entry_def['uid']} містить тільки байт типу. Пропускається.")
                continue
            type_str_from_map = type_map_ldk.get(file_type_val, 'bin')
            path_part_for_name = node_obj['path'].strip('/').replace('/', '_')
            if path_part_for_name:
                path_part_for_name = "_" + path_part_for_name
            contained_file_unique_name = f"{ldk_base_fn_for_contained}{path_part_for_name}_UID{entry_def.get('uid', 0):08X}.{type_str_from_map}"
            # --- ДІАГНОСТИКА ---
            print(f"LDK file UID={entry_def.get('uid')} type_byte=0x{file_type_val:02x} -> {type_str_from_map} ({contained_file_unique_name})")
            node_obj['files'].append({
                'name': contained_file_unique_name,
                'data_b64': base64.b64encode(actual_data_bytes).decode('ascii'),
                'type': type_str_from_map,
                'size': len(actual_data_bytes),
                'order': entry_def['_ix']
            })

        return node_obj

        def _tell(self):
            return self.rawoffs


        def _size(self):
            return self.rawsize

        def type(self):
            return self._file_type

        def data(self):
            return self.get_parsed_data()

        def get_parsed_data(self):
            output_data = {
                'ts': self.rawts, 'type': self._file_type,
                'path': self.path or self.rawname,
                'file': os.path.basename(self.path or self.rawname or "unknown_file"),
                'parse_successful': self.parse_successful
            }
            if self.parse_successful:
                if self._file_type == 'wpt':
                    output_data.update({'meta': self.data_parsed.get('meta'), 'location': self.data_parsed.get('location')})
                elif self._file_type in ['set', 'rte']:
                    output_data.update({'meta': self.data_parsed.get('meta'), 'waypoints': self.data_parsed.get('waypoints')})
                elif self._file_type == 'are':
                    output_data.update({'meta': self.data_parsed.get('meta'), 'locations': self.data_parsed.get('locations')})
                elif self._file_type == 'trk':
                    output_data.update({'meta': self.data_parsed.get('meta'),
                                       'waypoints': self.data_parsed.get('waypoints'),
                                       'segments': self.data_parsed.get('segments')})
                elif self._file_type == 'ldk':
                    output_data['root'] = self.data_parsed.get('root')
                elif self._file_type == 'bin':
                    output_data['raw_content_b64'] = self.data_parsed.get('raw_content_b64')
            return output_data  
    
class Main:
    """Головний клас програми з GUI для пакетної конвертації та обробки геоданих."""

    MAX_FILES: int = 100
    CSV_CHUNK_SIZE: int = 2000

    def __init__(self):
        self.program_version: str = "8.7.1_ukr_comments"
        self.empty: str = "Не вибрано"
        self.file_ext: Optional[str] = None
        self.file_name: Optional[str] = None

        self.list_of_formats = [".geojson", ".kml", ".kmz", ".gpx", ".xlsx", ".csv", ".csv(макет)"]
        self.supported_read_formats = [".kml", ".kmz", ".kme", ".gpx", ".xlsx", ".csv", ".scene", ".wpt", ".set",
                                       ".rte", ".are", ".trk", ".ldk"]
        self.numerations = ["За найближчими сусідами", "За змійкою", "За відстаню від кута", "За відстаню від границі",
                            "За випадковістю"]
        self.translations = ["Не повертати", "На 90 градусів", "На 180 градусів", "На 270 градусів"]

        self.colors = {
            "Red": "#f44336", "Pink": "#e91e63", "Purple": "#9c27b0", "DeepPurple": "#673ab7",
            "Indigo": "#3f51b5", "Blue": "#2196f3", "Cyan": "#00bcd4", "Teal": "#009688",
            "Green": "#4caf50", "LightGreen": "#8bc34a", "Lime": "#cddc39", "Yellow": "#ffeb3b",
            "Amber": "#ffc107", "Orange": "#ff9800", "DeepOrange": "#ff5722", "Brown": "#795548",
            "BlueGrey": "#607d8b", "Black": "#010101", "White": "#ffffff"
        }
        self.color_options = ["Без змін"] + list(self.colors.keys())

        self.colors_en_ua = {
            "Red": "Червоний", "Pink": "Рожевий", "Purple": "Фіолетовий", "DeepPurple": "Темно-фіолетовий",
            "Indigo": "Індиго", "Blue": "Синій", "Cyan": "Блакитний", "Teal": "Бірюзовий",
            "Green": "Зелений", "LightGreen": "Салатовий", "Lime": "Лаймовий", "Yellow": "Жовтий",
            "Amber": "Бурштиновий", "Orange": "Помаранчевий", "DeepOrange": "Насичено-помаранчевий",
            "Brown": "Коричневий", "BlueGrey": "Синьо-сірий", "Black": "Чорний", "White": "Білий"
        }

        self.color_keyword_map = {
            "червоний": "Red", "рожевий": "Pink", "фіолетовий": "Purple", "темно-фіолетовий": "DeepPurple",
            "індиго": "Indigo", "синій": "Blue", "блакитний": "Cyan", "бірюзовий": "Teal",
            "зелений": "Green", "салатовий": "LightGreen", "лаймовий": "Lime", "жовтий": "Yellow",
            "бурштиновий": "Amber", "помаранчевий": "Orange", "насичено-помаранчевий": "DeepOrange",
            "коричневий": "Brown", "синьо-сірий": "BlueGrey", "чорний": "Black", "білий": "White", "голубий": "Cyan",
            "красный": "Red", "розовый": "Pink", "фиолетовый": "Purple",
            "синий": "Blue", "зеленый": "Green", "желтый": "Yellow", "оранжевый": "Orange",
            "коричневый": "Brown", "черный": "Black", "белый": "White", "голубой": "Cyan"
        }

        self._palette_rgb = {
            name: (int(hx[1:3], 16), int(hx[3:5], 16), int(hx[5:7], 16))
            for name, hx in self.colors.items()
        }
        self.file_list: List[Dict[str, Any]] = []
        self.list_is_visible: bool = False
        self.output_directory_path: str = self.empty

        self.font = ("Courier New", 11, "bold")
        self.C_BACKGROUND, self.C_SIDEBAR, self.C_BUTTON, self.C_BUTTON_HOVER, self.C_TEXT = "#2B2B2B", "#3C3C3C", "#556B2F", "#6B8E23", "#F5F5F5"
        self.C_ACCENT_SUCCESS, self.C_ACCENT_DONE, self.C_STATUS_DEFAULT, self.C_ACCENT_ERROR = "#6B8E23", "#FFBF00", "#4F4F4F", "#8B0000"

        self.main_window = tk.Tk()
        self.names_agree = tk.BooleanVar(value=False)
        self.exceptions_agree = tk.BooleanVar(value=False)
        self.chosen_numeration = tk.StringVar(value="За найближчими сусідами")
        self.chosen_translation = tk.StringVar(value="Не повертати")

        self.main_window.title(f"Nexus v{self.program_version}")
        self.main_window.configure(background=self.C_BACKGROUND)
        self.main_window.minsize(450, 120)
        self.main_window.geometry("450x120")
        try:
            base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
            icon_path = os.path.join(base_path, 'nexus.ico')
            if os.path.exists(icon_path): self.main_window.iconbitmap(icon_path)
        except Exception as e:
            print(f"Попередження: не вдалося завантажити іконку: {e}")
        self.main_window.protocol("WM_DELETE_WINDOW", self.exit)
        self.main_window.resizable(True, True)

        self._configure_styles()
        self._build_main_ui()

        self.input_file_path: Optional[str] = None
        self.output_directory_path: str = self.empty
        
    def _configure_styles(self):
        style = ttk.Style(self.main_window)
        style.theme_use('clam')
        style.configure("TFrame", background=self.C_BACKGROUND)
        style.configure("Side.TFrame", background=self.C_SIDEBAR)
        style.configure("List.TFrame", background=self.C_SIDEBAR)
        style.configure('Icon.TButton', padding=5, borderwidth=0, relief='flat', background=self.C_BUTTON,
                        foreground=self.C_TEXT, font=self.font)
        style.map('Icon.TButton', background=[('active', self.C_BUTTON_HOVER)], foreground=[('active', self.C_TEXT)])
        style.configure('Remove.TButton', background=self.C_SIDEBAR, foreground="#FF6347",
                        font=("Courier New", 10, "bold"), relief='flat', borderwidth=0)
        style.map('Remove.TButton', background=[('active', "#4a4a4a")])
        style.configure("Toplevel", background=self.C_BACKGROUND)
        style.configure("TCheckbutton", background=self.C_BACKGROUND, foreground=self.C_TEXT, font=self.font,
                        indicatorcolor=self.C_TEXT, selectcolor=self.C_BUTTON_HOVER)
        style.map("TCheckbutton", background=[('active', self.C_BACKGROUND)])
        style.configure("TLabel", background=self.C_BACKGROUND, foreground=self.C_TEXT, font=self.font)
        style.configure("List.TLabel", background=self.C_SIDEBAR, foreground=self.C_TEXT, font=("Courier New", 9))
        style.configure("Dark.TEntry", fieldbackground="#4F4F4F", foreground=self.C_TEXT, insertcolor=self.C_TEXT,
                        bordercolor=self.C_SIDEBAR, font=("Courier New", 9))
        style.configure("TMenubutton", background="#4F4F4F", foreground=self.C_TEXT, font=("Courier New", 9),
                        borderwidth=1, relief='raised', arrowcolor=self.C_TEXT)
        style.map("TMenubutton", background=[('active', "#646464")])

    def run(self):
        self.main_window.mainloop()

    def exit(self):
        if messagebox.askokcancel("Вихід", "Ви впевнені, що хочете вийти?"):
            self.main_window.destroy()

    def _build_main_ui(self):
        self.main_window.rowconfigure(0, weight=0)
        self.main_window.rowconfigure(1, weight=1)
        self.main_window.columnconfigure(0, weight=1)
        top_container = ttk.Frame(self.main_window)
        top_container.grid(row=0, column=0, sticky="ew", pady=(5, 0))
        top_container.columnconfigure(0, weight=0)
        top_container.columnconfigure(1, weight=1)
        top_container.columnconfigure(2, weight=0)
        left_sidebar = ttk.Frame(top_container, width=50, style="Side.TFrame")
        left_sidebar.grid(row=0, column=0, sticky="ns", padx=(5, 2))
        btn_lightbulb = ttk.Button(left_sidebar, text="i", style='Icon.TButton', command=self.show_info, width=2)
        btn_lightbulb.pack(pady=(5, 5), padx=5, fill='x')
        Tooltip(btn_lightbulb, "Про програму", background=self.C_SIDEBAR, foreground=self.C_TEXT)
        btn_settings = ttk.Button(left_sidebar, text="S", style='Icon.TButton', command=self.open_numeration_settings,
                                  width=2)
        btn_settings.pack(pady=5, padx=5, fill='x')
        Tooltip(btn_settings, "Налаштування нумерації", background=self.C_SIDEBAR, foreground=self.C_TEXT)
        center_frame = ttk.Frame(top_container)
        center_frame.grid(row=0, column=1, sticky="nsew", padx=0, pady=0)
        self.status_label = ttk.Label(center_frame, anchor="center", font=("Courier New", 14, "bold"),
                                      foreground=self.C_TEXT, relief='flat', padding=(0, 10))
        self.status_label.pack(fill="both", expand=True)
        self._update_status("ДОДАЙТЕ ФАЙЛИ", self.C_STATUS_DEFAULT)
        right_sidebar = ttk.Frame(top_container, width=50, style="Side.TFrame")
        right_sidebar.grid(row=0, column=2, sticky="ns", padx=(2, 5))
        self.btn_open_file = ttk.Button(right_sidebar, text="F", style='Icon.TButton',
                                        command=self.add_files_to_list, width=2)
        self.btn_open_file.pack(pady=(5, 5), padx=5, fill='x')
        Tooltip(self.btn_open_file, "Додати файли", background=self.C_SIDEBAR, foreground=self.C_TEXT)
        self.play_button = ttk.Button(right_sidebar, text="▶", style='Icon.TButton', command=self.start_convertion,
                                      state="disabled", width=2)
        self.play_button.pack(pady=5, padx=5, fill='x')
        Tooltip(self.play_button, "Конвертувати все", background=self.C_SIDEBAR, foreground=self.C_TEXT)
        self.list_container = ttk.Frame(self.main_window, style="List.TFrame")
        self.canvas = tk.Canvas(self.list_container, bg=self.C_SIDEBAR, highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.list_container, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas, style="List.TFrame")
        self.scrollable_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas_window = self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=scrollbar.set)
        self.canvas.pack(side="left", fill="both", expand=True, padx=(1, 0), pady=(1, 1))
        scrollbar.pack(side="right", fill="y", padx=(0, 1), pady=(1, 1))
        self.canvas.bind("<Configure>", self._on_canvas_configure)

    def _on_canvas_configure(self, event):
        self.canvas.itemconfig(self.canvas_window, width=event.width)
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _redraw_file_list(self):
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()
        for i, file_data in enumerate(self.file_list):
            item_frame = ttk.Frame(self.scrollable_frame, style="List.TFrame", padding=(5, 2))
            item_frame.pack(fill='x', expand=True)
            label_text = f"{i + 1}. {file_data['base_name']}"
            if len(label_text) > 40: label_text = label_text[:37] + "..."
            label = ttk.Label(item_frame, text=label_text, style="List.TLabel", anchor='w')
            label.pack(side='left', fill='x', expand=True, padx=(0, 5))
            format_mb = ttk.Menubutton(item_frame, text=file_data['format_var'].get(), style="TMenubutton", width=10)
            format_menu_tk = tk.Menu(format_mb, tearoff=0, bg=self.C_SIDEBAR, fg=self.C_TEXT,
                                     activebackground=self.C_BUTTON_HOVER)
            for fmt_option in self.list_of_formats:
                format_menu_tk.add_radiobutton(label=fmt_option, variable=file_data['format_var'], value=fmt_option,
                                               command=lambda var=file_data['format_var'], button=format_mb,
                                                              val=fmt_option: self._update_menubutton_text(var, button,
                                                                                                          val))
            format_mb['menu'] = format_menu_tk
            format_mb.pack(side='left', padx=3)
            color_mb = ttk.Menubutton(item_frame,
                                      text=self.colors_en_ua.get(file_data['color_var'].get(),
                                                                file_data['color_var'].get()), style="TMenubutton",
                                      width=12)
            color_menu_tk = tk.Menu(color_mb, tearoff=0, bg=self.C_SIDEBAR, fg=self.C_TEXT,
                                    activebackground=self.C_BUTTON_HOVER)
            for color_option in self.color_options:
                disp_name = self.colors_en_ua.get(color_option, color_option)
                color_menu_tk.add_radiobutton(label=disp_name, variable=file_data['color_var'], value=color_option,
                                              command=lambda var=file_data['color_var'], button=color_mb,
                                                             val_en=color_option: self._update_menubutton_text(var,
                                                                                                               button,
                                                                                                               val_en))
            color_mb['menu'] = color_menu_tk
            color_mb.pack(side='left', padx=3)
            remove_btn = ttk.Button(item_frame, text="X", style='Remove.TButton', width=2,
                                    command=lambda fd=file_data: self._remove_file(fd))
            remove_btn.pack(side='left', padx=(3, 0))
        if not self.file_list:
            self._update_status("ДОДАЙТЕ ФАЙЛИ", self.C_STATUS_DEFAULT)
            self.play_button.config(state="disabled")
            if self.list_is_visible:
                self.list_container.grid_forget()
                self.list_is_visible = False
                self.main_window.geometry("450x120")
        else:
            status_text = f"ГОТОВО: {len(self.file_list)} ФАЙЛ(ІВ)"
            if len(self.file_list) == 1:
                status_text = f"ГОТОВО: {len(self.file_list)} ФАЙЛ"
            elif 2 <= len(self.file_list) <= 4:
                status_text = f"ГОТОВО: {len(self.file_list)} ФАЙЛИ"
            self._update_status(status_text, self.C_ACCENT_SUCCESS)
            self.play_button.config(state="normal")
        self.scrollable_frame.update_idletasks()
        self.canvas.config(scrollregion=self.canvas.bbox("all"))
        self.canvas.itemconfig(self.canvas_window, width=self.canvas.winfo_width())

    def _update_menubutton_text(self, var, menubutton, value_english):
        var.set(value_english)
        display_text = self.colors_en_ua.get(value_english, value_english)
        if not display_text: display_text = value_english
        menubutton.config(text=display_text)

    def _remove_file(self, file_to_remove):
        self.file_list.remove(file_to_remove)
        if not self.file_list and self.list_is_visible:
            self.list_container.grid_forget()
            self.list_is_visible = False
            self.main_window.geometry("450x120")
        self._redraw_file_list()

    def add_files_to_list(self):
        file_types = [("Підтримувані файли", " ".join(f"*{ext}" for ext in self.supported_read_formats)),
                      ("AlpineQuest файли", ".wpt .set .rte .are .trk .ldk"), ("KML/KMZ/KME", ".kml .kmz .kme"),
                      ("GPS Exchange", ".gpx"), ("Excel", ".xlsx"), ("CSV", ".csv"), ("SCENE JSON", ".scene"),
                      ("Всі файли", "*.*")]
        paths = filedialog.askopenfilenames(filetypes=file_types, title="Виберіть файли для конвертації")
        new_files_added = False
        if paths:
            for path in paths:
                if any(f['full_path'] == path for f in self.file_list):
                    self._update_status(f"Файл вже у списку: {os.path.basename(path)}", warning=True)
                    continue
                if len(self.file_list) >= self.MAX_FILES:
                    messagebox.showwarning("Ліміт файлів",
                                           f"Максимальна кількість файлів у списку ({self.MAX_FILES}) досягнута.")
                    break
                base_name = os.path.basename(path)
                file_ext = os.path.splitext(base_name)[1].lower()
                if file_ext not in self.supported_read_formats:
                    messagebox.showwarning("Формат не підтримується",
                                           f"Програма не може імпортувати дані з файлів формату '{file_ext}'.")
                    continue
                default_export_format = ".kml"
                if file_ext in [".wpt", ".set", ".rte", ".are", ".trk", ".ldk", ".kmz", ".kme"]:
                    default_export_format = ".kml"
                elif file_ext == ".gpx":
                    default_export_format = ".gpx"
                elif file_ext == ".xlsx":
                    default_export_format = ".xlsx"
                elif file_ext == ".csv":
                    default_export_format = ".csv"
                elif file_ext == ".scene":
                    default_export_format = ".geojson"
                file_data = {"full_path": path, "base_name": base_name,
                             "format_var": tk.StringVar(value=default_export_format),
                             "color_var": tk.StringVar(value=self.color_options[0])}
                self.file_list.append(file_data)
                new_files_added = True
            if new_files_added and not self.list_is_visible:
                self.list_container.grid(row=1, column=0, columnspan=3, sticky="nsew", padx=5, pady=(2, 5))
                self.main_window.rowconfigure(1, weight=3)
                self.list_is_visible = True
                self.main_window.geometry("650x400")
            if new_files_added:
                self._redraw_file_list()
            elif not self.file_list:
                self._update_status("ФАЙЛИ НЕ ДОДАНО", self.C_STATUS_DEFAULT)

    def _update_status(self, text, color=None, error=False, warning=False):
        if error:
            final_color = self.C_ACCENT_ERROR
        elif warning:
            final_color = self.C_ACCENT_DONE
        elif color:
            final_color = color
        else:
            final_color = self.C_TEXT

        if not error and not warning and not color and self.status_label.cget("background") == self.C_STATUS_DEFAULT:
            current_bg = self.C_STATUS_DEFAULT
        elif not error and not warning and not color:
            current_bg = self.C_STATUS_DEFAULT if self.status_label.cget(
                "background") == self.C_STATUS_DEFAULT else self.C_BACKGROUND
        else:
            current_bg = final_color

        self.status_label.config(text=text.upper(), background=current_bg, foreground=self.C_TEXT)
        if self.main_window.winfo_exists():
            self.main_window.update_idletasks()

    def _get_chunked_save_path(self, base_save_path, chunk_index):
        if chunk_index == 0:
            return base_save_path
        directory, filename = os.path.split(base_save_path)
        name_part, ext_part = os.path.splitext(filename)
        name_part = re.sub(r'\(\d+\)$', '', name_part).strip()
        new_filename = f"{name_part}({chunk_index + 1}){ext_part}"
        return os.path.join(directory, new_filename)
        
    def show_info(self):
        messagebox.showinfo(
            "Про програму",
            f"Nexus v{self.program_version}\n"
            "Програма для пакетної конвертації та обробки геоданих.\n\n"
            f"Підтримувані формати для читання:\n{', '.join(self.supported_read_formats)}\n\n"
            f"Підтримувані формати для запису:\n{', '.join(fmt for fmt in self.list_of_formats if fmt != '.csv(макет)')}"
        )
        
    def open_numeration_settings(self):
        settings_win = tk.Toplevel(self.main_window)
        settings_win.title("Налаштування нумерації")
        settings_win.configure(background=self.C_BACKGROUND)
        settings_win.transient(self.main_window)
        settings_win.grab_set()
        settings_win.resizable(False, False)

        main_frame = ttk.Frame(settings_win, padding=15)
        main_frame.pack(fill="both", expand=True)

        ttk.Checkbutton(main_frame, text="Увімкнути нумерацію точок", variable=self.names_agree).pack(anchor="w", pady=(0, 10))

        numeration_frame = ttk.LabelFrame(main_frame, text="Параметри нумерації", style="TFrame", padding=10)
        numeration_frame.pack(fill="x", expand=True, pady=5)
        ttk.Label(numeration_frame, text="Спосіб нумерації:").pack(anchor="w")
        numeration_combo = ttk.Combobox(
            numeration_frame, textvariable=self.chosen_numeration,
            values=self.numerations, state="readonly", font=("Courier New", 9)
        )
        numeration_combo.pack(fill="x", pady=(2, 8))
        numeration_combo.set(self.chosen_numeration.get())

        ttk.Label(numeration_frame, text="Поворот осі сортування:").pack(anchor="w")
        translation_combo = ttk.Combobox(
            numeration_frame, textvariable=self.chosen_translation,
            values=self.translations, state="readonly", font=("Courier New", 9)
        )
        translation_combo.pack(fill="x", pady=(2, 8))
        translation_combo.set(self.chosen_translation.get())

        ttk.Checkbutton(
            numeration_frame,
            text="Виключити номери (30-40, 500-510)",
            variable=self.exceptions_agree
        ).pack(anchor="w", pady=5)

        ttk.Button(main_frame, text="Закрити", command=settings_win.destroy, style="Icon.TButton").pack(pady=(15, 0))

        settings_win.update_idletasks()
        main_x, main_y = self.main_window.winfo_x(), self.main_window.winfo_y()
        main_w, main_h = self.main_window.winfo_width(), self.main_window.winfo_height()
        win_w, win_h = settings_win.winfo_width(), settings_win.winfo_height()
        x = main_x + (main_w // 2) - (win_w // 2)
        y = main_y + (main_h // 2) - (win_h // 2)
        settings_win.geometry(f"+{x}+{y}")
        settings_win.focus_set()
        
    def _process_data(self, file_content, color_override_english_name):
        if not file_content:
            return None
        content = copy.deepcopy(file_content)

        if color_override_english_name != self.color_options[0]:
            for item in content:
                item["color"] = color_override_english_name
                if 'milgeo:meta:color' in item:
                    item['milgeo:meta:color'] = color_override_english_name

        if self.names_agree.get():
            points_to_numerate = [item for item in content if item.get('geometry_type', '').lower() == 'point']
            other_items = [item for item in content if item.get('geometry_type', '').lower() != 'point']

            if points_to_numerate:
                numerated_points = self._apply_selected_numeration(points_to_numerate)
                content = numerated_points + other_items

        return content
        
    def start_convertion(self):
        if not self.file_list:
            messagebox.showwarning("Увага", "Список файлів для конвертації порожній.")
            return
        readers = {
            ".kml": self.read_kml,
            ".kme": self.read_kml,
            ".kmz": self.read_kmz,
            ".gpx": self.read_gpx,
            ".xlsx": self.read_xlsx,
            ".csv": self.read_csv,
            ".scene": self.read_scene,
            ".geojson": self.read_geojson,
            ".wpt": self.read_wpt,
            ".set": self.read_set,
            ".rte": self.read_rte,
            ".are": self.read_are,
            ".trk": self.read_trk,
            ".ldk": self.read_ldk,
        }

        writers = {
            ".kml": self.create_kml,
            ".kme": self.create_kml,
            ".kmz": self.create_kmz,
            ".gpx": self.create_gpx,
            ".xlsx": self.create_xlsx,
            ".csv": self.create_csv,
            ".csv(макет)": self.create_csv,
            ".geojson": self.create_geojson,
            ".scene": self.create_scene,
        }
        total_files = len(self.file_list)
        conversion_successful_count = 0

        for i, file_data in enumerate(self.file_list):
            current_file_basename = file_data['base_name']
            self._update_status(f"ФАЙЛ {i + 1}/{total_files}: {current_file_basename}", self.C_BUTTON_HOVER)
            self.main_window.update_idletasks()
            try:
                input_path = file_data['full_path']
                self.file_name, self.file_ext = os.path.splitext(os.path.basename(input_path))
                self.file_ext = self.file_ext.lower()
                reader_func = readers.get(self.file_ext)
                if not reader_func:
                    self._update_status(f"Непідтримуваний формат для читання: {self.file_ext}", error=True)
                    continue
                file_content = reader_func(input_path)
                if file_content is None or not file_content:
                    messagebox.showwarning("Увага", f"У файлі {current_file_basename} не знайдено даних або сталася помилка читання. Файл пропущено.")
                    self._update_status(f"ПОМИЛКА ЧИТАННЯ: {current_file_basename}", warning=True)
                    continue
                processed_content = self._process_data(file_content, file_data['color_var'].get())
                if processed_content is None:
                    self._update_status(f"Помилка обробки даних: {current_file_basename}", error=True)
                    continue
                output_format = file_data['format_var'].get().lower()
                writer_func = writers.get(output_format)
                if not writer_func:
                    messagebox.showerror("Формат не підтримується", f"Конвертація у формат '{output_format}' не підтримується.")
                    continue
                clean_base_name = re.sub(r'\(\d+\)$', '', self.file_name).strip()
                suggested_name = f"new_{clean_base_name}{output_format.replace('(макет)', '')}"
                if self.output_directory_path == self.empty or not os.path.isdir(self.output_directory_path):
                    self.output_directory_path = os.path.dirname(input_path)
                save_path = filedialog.asksaveasfilename(
                    initialdir=self.output_directory_path,
                    initialfile=suggested_name,
                    defaultextension=output_format.replace('(макет)', ''),
                    filetypes=[(f"{output_format.upper()} Files", f"*{output_format.replace('(макет)', '')}"),
                               ("All Files", "*.*")],
                    title=f"Зберегти конвертований файл для {current_file_basename}"
                )
                if not save_path:
                    self._update_status(f"СКАСОВАНО: {current_file_basename}", warning=True)
                    continue
                self.output_directory_path = os.path.dirname(save_path)
                success = writer_func(processed_content, save_path)
                if success:
                    if output_format not in ['.csv', '.csv(макет)']:
                        self._update_status(f"ЗБЕРЕЖЕНО: {os.path.basename(save_path)}", self.C_ACCENT_SUCCESS)
                    conversion_successful_count += 1
            except NotImplementedError as e:
                messagebox.showerror("Не реалізовано", str(e))
                self._update_status(f"ПОМИЛКА ФОРМАТУ: {current_file_basename}", error=True)
            except ValueError as e:
                messagebox.showerror("Помилка даних", f"Проблема з даними у файлі {current_file_basename}: {e}")
                self._update_status(f"ПОМИЛКА ДАНИХ: {current_file_basename}", error=True)
            except Exception as e:
                messagebox.showerror("Критична помилка", f"Не вдалося конвертувати файл {current_file_basename}:\n\n{type(e).__name__}: {e}\n\nПеревірте консоль для деталей.")
                self._update_status(f"КРИТИЧНА ПОМИЛКА: {current_file_basename}", self.C_ACCENT_ERROR)
                import traceback
                traceback.print_exc()
                continue

        if conversion_successful_count == total_files and total_files > 0:
            final_message = f"УСПІШНО ЗАВЕРШЕНО: {conversion_successful_count}/{total_files} ФАЙЛ(ІВ)"
            final_color = self.C_ACCENT_DONE
        elif conversion_successful_count > 0:
            final_message = f"ЗАВЕРШЕНО З ПОМИЛКАМИ: {conversion_successful_count}/{total_files} УСПІШНО"
            final_color = self.C_ACCENT_DONE
        elif total_files > 0:
            final_message = f"ПОМИЛКА: 0/{total_files} ФАЙЛІВ КОНВЕРТОВАНО"
            final_color = self.C_ACCENT_ERROR
        else:
            final_message = "СПИСОК ПОРОЖНІЙ"
            final_color = self.C_STATUS_DEFAULT

        self._update_status(final_message, final_color)
        if total_files > 0:
            messagebox.showinfo("Завершено", f"Пакетна конвертація завершена.\nУспішно: {conversion_successful_count} з {total_files}.")

    def _normalize_apq_data(self, apq_parsed_data, file_path_for_log="", expand_lines_to_points=False):
        normalized_content = []
        if not apq_parsed_data or not isinstance(apq_parsed_data, dict) or not apq_parsed_data.get('parse_successful'):
            self._update_status(f"APQ парсер не повернув успішних даних для {file_path_for_log}", warning=True)
            return normalized_content

        apq_type = apq_parsed_data.get('type')
        global_meta = apq_parsed_data.get('meta', {})
        file_basename = os.path.basename(file_path_for_log)

        def _create_point_dict(loc_data, item_meta_data, default_name_prefix="Точка", item_idx=0,
                               apq_source_file_type_for_item=None,
                               source_file_global_meta_for_item=None):
            if not loc_data or loc_data.get('lon') is None or loc_data.get('lat') is None:
                self._update_status(f"Увага: Пропущено точку (відсутні координати) у {file_basename}", warning=True)
                return None
            point_lon, point_lat = loc_data['lon'], loc_data['lat']

            effective_meta = source_file_global_meta_for_item.copy() if source_file_global_meta_for_item else {}
            effective_meta.update(item_meta_data)

            final_name = effective_meta.get('name', f"{default_name_prefix}_{item_idx + 1}")
            point_type_val = effective_meta.get('sym', effective_meta.get('icon', 'Landmark'))
            point_color_str = effective_meta.get('color', "White")
            description_val = effective_meta.get('comment', effective_meta.get('description', ''))

            extra_desc_parts = []
            if loc_data.get('alt') is not None: extra_desc_parts.append(f"Висота: {loc_data['alt']:.1f}м")
            if loc_data.get('ts') is not None:
                try:
                    dt_obj = datetime.fromtimestamp(loc_data['ts'], timezone.utc)
                    extra_desc_parts.append(f"Час: {dt_obj.strftime('%Y-%m-%d %H:%M:%S %Z')}")
                except:
                    pass
            if loc_data.get('acc') is not None: extra_desc_parts.append(f"Точність: {loc_data['acc']:.1f}м")

            full_description = str(description_val) if description_val else ""
            if extra_desc_parts:
                full_description = (full_description + " | " if full_description else "") + "; ".join(extra_desc_parts)

            entry = {
                "name": final_name, "lat": point_lat, "lon": point_lon,
                "type": point_type_val,
                "description": full_description if full_description else None,
                "geometry_type": "Point",
                "color": point_color_str,
                "original_location_data": loc_data,
                "apq_original_type": apq_source_file_type_for_item,
                'milgeo:meta:name': final_name,
                'milgeo:meta:color': point_color_str,
                'milgeo:meta:desc': description_val,
                'milgeo:meta:creator': effective_meta.get('creator'),
                'milgeo:meta:creator_url': effective_meta.get('creator_url'),
                'milgeo:meta:sidc': effective_meta.get('sidc', global_meta.get('sidc'))
            }
            return entry

        # --- WPT
        if apq_type == 'wpt':
            loc = apq_parsed_data.get('location')
            point = _create_point_dict(loc, global_meta, "Waypoint",
                                       apq_source_file_type_for_item='wpt',
                                       source_file_global_meta_for_item=global_meta)
            if point: normalized_content.append(point)

        # --- SET, RTE
        elif apq_type in ['set', 'rte']:
            waypoints_list = apq_parsed_data.get('waypoints', [])
            default_prefix = global_meta.get('name', apq_type.upper())
            for idx, wpt_entry in enumerate(waypoints_list):
                point = _create_point_dict(
                    wpt_entry.get('location'), wpt_entry.get('meta', {}),
                    default_prefix, idx, apq_source_file_type_for_item=apq_type,
                    source_file_global_meta_for_item=global_meta
                )
                if point: normalized_content.append(point)

        # --- ARE (полігон)
        elif apq_type == 'are':
            locations_list = apq_parsed_data.get('locations', [])
            area_name = global_meta.get('name', 'Area')
            area_color_str = global_meta.get('color', "Blue")
            area_description = global_meta.get('comment', global_meta.get('description', ''))
            area_points_data_for_polygon = [loc for loc in locations_list if loc and loc.get('lon') is not None]

            if len(area_points_data_for_polygon) >= 3:
                poly_item = {
                    'name': area_name, 'type': 'Area', 'geometry_type': 'Polygon',
                    'points_data': area_points_data_for_polygon,
                    'apq_original_type': 'are',
                    'color': area_color_str,
                    'description': area_description,
                    'milgeo:meta:name': area_name, 'milgeo:meta:color': area_color_str,
                    'milgeo:meta:desc': area_description,
                    'milgeo:meta:creator': global_meta.get('creator'),
                    'milgeo:meta:creator_url': global_meta.get('creator_url'),
                    'milgeo:meta:sidc': global_meta.get('sidc')
                }
                normalized_content.append(poly_item)

        # --- TRK (POI + segments)
        elif apq_type == 'trk':
            track_default_name = global_meta.get('name', 'Track')
            # POI точки
            for idx, poi_entry in enumerate(apq_parsed_data.get('waypoints', [])):
                point = _create_point_dict(
                    poi_entry.get('location'), poi_entry.get('meta', {}),
                    f"{track_default_name}_POI", idx,
                    apq_source_file_type_for_item='trk_poi',
                    source_file_global_meta_for_item=global_meta
                )
                if point: normalized_content.append(point)
            # Сегменти (лінії)
            segments_data = apq_parsed_data.get('segments', [])
            for seg_idx, segment_item in enumerate(segments_data):
                seg_locs_data = segment_item.get('locations', [])
                seg_meta_data = segment_item.get('meta', {})
                if len(seg_locs_data) < 2: continue

                effective_seg_meta = global_meta.copy()
                effective_seg_meta.update(seg_meta_data)

                segment_name = effective_seg_meta.get('name', f"{track_default_name}_Segment_{seg_idx + 1}")
                segment_color_str = effective_seg_meta.get('color', global_meta.get('color', "Red"))
                segment_description = effective_seg_meta.get('comment', effective_seg_meta.get('description', ''))
                segment_points_for_line = [loc for loc in seg_locs_data if loc and loc.get('lon') is not None]

                if len(segment_points_for_line) >= 2:
                    line_item = {
                        'name': segment_name, 'geometry_type': 'LineString',
                        'points_data': segment_points_for_line,
                        'apq_original_type': 'trk',
                        'color': segment_color_str,
                        'description': segment_description,
                        'milgeo:meta:name': segment_name, 'milgeo:meta:color': segment_color_str,
                        'milgeo:meta:desc': segment_description,
                        'milgeo:meta:creator': global_meta.get('creator'),
                        'milgeo:meta:creator_url': global_meta.get('creator_url'),
                        'milgeo:meta:sidc': effective_seg_meta.get('sidc', global_meta.get('sidc'))
                    }
                    normalized_content.append(line_item)

        # --- Розгортання ліній/полігонів у точки (для CSV, якщо потрібно)
        if expand_lines_to_points:
            expanded_points = []
            for item in normalized_content:
                if item.get('geometry_type') in ('LineString', 'Polygon'):
                    for idx, pt in enumerate(item.get('points_data', [])):
                        if pt.get('lon') is not None and pt.get('lat') is not None:
                            pt_entry = {
                                "name": f"{item.get('name')}_pt{idx+1}",
                                "lat": pt.get('lat'), "lon": pt.get('lon'),
                                "type": item.get('type') if 'type' in item else item.get('geometry_type'),
                                "description": item.get('description'),
                                "geometry_type": "Point",
                                "color": item.get('color'),
                                "original_location_data": pt,
                                "apq_original_type": item.get('apq_original_type'),
                                'milgeo:meta:name': item.get('name'),
                                'milgeo:meta:color': item.get('color'),
                                'milgeo:meta:desc': item.get('description'),
                                'milgeo:meta:sidc': item.get('milgeo:meta:sidc')
                            }
                            expanded_points.append(pt_entry)
            normalized_content += expanded_points

        # Діагностика
        print(f"[normalize_apq_data] {apq_type}: "
              f"{sum(1 for i in normalized_content if i['geometry_type']=='Point')} pts, "
              f"{sum(1 for i in normalized_content if i['geometry_type']=='LineString')} lines, "
              f"{sum(1 for i in normalized_content if i['geometry_type']=='Polygon')} polygons")

        if not normalized_content and apq_type not in ['ldk', 'bin']:
            self._update_status(f"Увага: Не знайдено даних для нормалізації у {file_basename} (тип {apq_type})",
                                warning=True)
        return normalized_content

    def _read_specific_file(self, file_path_to_read, expected_file_extension):
        self.input_file_path = file_path_to_read
        content_list = []
        file_basename_log = os.path.basename(file_path_to_read)
        try:
            apq_parser_instance = ApqFile(path=file_path_to_read, verbosity=0, gui_logger_func=self._update_status)

            if not apq_parser_instance.parse_successful:
                self._update_status(f"Помилка парсингу APQ файлу: {file_basename_log}", error=True)
                return None

            apq_data_structure = apq_parser_instance.data()
            content_list = self._normalize_apq_data(apq_data_structure, file_path_to_read)

        except ValueError as e:
            self._update_status(f"Помилка ініціалізації парсера для {file_path_to_read}: {e}", error=True)
            return None
        except Exception as e:
            self._update_status(f"Загальна помилка парсингу APQ {file_path_to_read}: {e}", error=True)
            print(f"Загальний виняток APQ ({file_path_to_read}): {type(e).__name__}: {e}")
            return None

        return content_list if content_list else None

    def read_wpt(self, path):
        return self._read_specific_file(path, ".wpt")

    def read_set(self, path):
        return self._read_specific_file(path, ".set")

    def read_rte(self, path):
        return self._read_specific_file(path, ".rte")

    def read_are(self, path):
        return self._read_specific_file(path, ".are")

    def read_trk(self, path):
        return self._read_specific_file(path, ".trk")

    def read_ldk(self, path):
        """
        Читання LDK-файлу з повною рекурсією по всіх nodes/files, витягує всі точки, лінії, полігони.
        """
        self._update_status(f"Читання LDK: {os.path.basename(path)}...", self.C_BUTTON_HOVER)
        all_normalized_content = []
        stats = {'nodes': 0, 'files': 0, 'points': 0, 'lines': 0, 'polygons': 0}

        try:
            ldk_apq_file_instance = ApqFile(path=path, verbosity=0, gui_logger_func=self._update_status)
            if not ldk_apq_file_instance.parse_successful:
                messagebox.showerror("Помилка LDK", f"Не вдалося розпарсити LDK файл: {os.path.basename(path)}")
                return None

            parsed_ldk_root_data = ldk_apq_file_instance.data()

            def extract_and_normalize_from_ldk_node(node_data, parent_original_path, depth=0):
                if not node_data:
                    return

                stats['nodes'] += 1
                node_path = node_data.get('path', '')

                # Діагностика структури LDK
                print(f"{'  ' * depth}[LDK-NODE] path={node_path} files={len(node_data.get('files', []))} nodes={len(node_data.get('nodes', []))}")

                # Обробка всіх файлів (маршрути, треки, полігони, точки)
                for ldk_file_entry in node_data.get('files', []):
                    stats['files'] += 1
                    inner_file_type = ldk_file_entry.get('type')
                    inner_file_name = ldk_file_entry.get('name')

                    self._update_status(f"Обробка з LDK: {inner_file_name}", self.C_BUTTON_HOVER)

                    if inner_file_type == 'bin':
                        self._update_status(f".bin з LDK: {inner_file_name}, експорт не підтримується.", warning=True)
                        continue

                    try:
                        file_content_bytes = base64.b64decode(ldk_file_entry['data_b64'])
                    except Exception as e:
                        self._update_status(f"Не вдалося декодувати base64 для {inner_file_name}: {e}", error=True)
                        continue

                    try:
                        contained_apq = ApqFile(
                            rawdata=file_content_bytes,
                            file_type=inner_file_type,
                            rawname=inner_file_name,
                            rawts=parsed_ldk_root_data.get('ts', time.time()),
                            verbosity=0,
                            gui_logger_func=self._update_status
                        )
                        if contained_apq.parse_successful:
                            # ВАЖЛИВО! Розгортаємо всі лінії/полігони у точки для CSV (expand_lines_to_points=True)
                            normalized_data = self._normalize_apq_data(
                                contained_apq.data(), inner_file_name, expand_lines_to_points=True
                            )
                            if normalized_data:
                                for item_norm in normalized_data:
                                    item_norm['source_file'] = inner_file_name
                                    item_norm['ldk_parent'] = os.path.basename(parent_original_path)
                                    # Підрахунок типів для діагностики
                                    if item_norm.get('geometry_type') == 'Point':
                                        stats['points'] += 1
                                    elif item_norm.get('geometry_type') == 'LineString':
                                        stats['lines'] += 1
                                    elif item_norm.get('geometry_type') == 'Polygon':
                                        stats['polygons'] += 1
                                all_normalized_content.extend(normalized_data)
                            else:
                                print(f"{'  ' * depth}[LDK-FILE] {inner_file_name}: дані не нормалізовані")
                        else:
                            self._update_status(f"Помилка парсингу файлу з LDK: {inner_file_name}", warning=True)
                    except Exception as e:
                        self._update_status(f"Помилка обробки {inner_file_name} з LDK: {e}", error=True)
                        import traceback
                        traceback.print_exc()

                # Рекурсивно обробити всі дочірні вузли
                for child_node in node_data.get('nodes', []):
                    extract_and_normalize_from_ldk_node(child_node, parent_original_path, depth=depth+1)

            # Запуск рекурсії від root
            if parsed_ldk_root_data and parsed_ldk_root_data.get('root'):
                extract_and_normalize_from_ldk_node(parsed_ldk_root_data['root'], path, depth=0)
                print(f"[LDK-DONE] Всього вузлів: {stats['nodes']}, файлів: {stats['files']}, точок: {stats['points']}, ліній: {stats['lines']}, полігонів: {stats['polygons']}")
            else:
                messagebox.showwarning("Увага LDK", f"LDK файл {os.path.basename(path)} порожній або має невірну структуру.")
                return None

        except ValueError as e:
            messagebox.showerror("Помилка LDK", f"Не вдалося ініціалізувати парсер для LDK {os.path.basename(path)}: {e}")
            return None
        except Exception as e:
            messagebox.showerror("Помилка читання LDK", f"Не вдалося обробити файл {os.path.basename(path)}.\n{type(e).__name__}: {e}")
            import traceback
            traceback.print_exc()
            return None

        return all_normalized_content if all_normalized_content else None


    def _read_kml_from_content(self, kml_content, source_filename="KML"):
        result = []
        try:
            if kml_content.startswith('\ufeff'):
                kml_content = kml_content[1:]
            root = ET.fromstring(kml_content)
            ns_match = re.match(r'\{([^}]+)\}', root.tag)
            ns_uri = ns_match.group(1) if ns_match else 'http://www.opengis.net/kml/2.2'
            ns = {'kml': ns_uri} if root.tag.startswith(f'{{{ns_uri}}}') else {}

            def find_tag(element, tag_name):
                if ns: return element.find(f'.//kml:{tag_name}', namespaces=ns)
                return element.find(f'.//{tag_name}')

            def findall_tags(element, tag_name):
                if ns: return element.findall(f'.//kml:{tag_name}', namespaces=ns)
                return element.findall(f'.//{tag_name}')

            for placemark in findall_tags(root, 'Placemark'):
                name_tag = find_tag(placemark, 'name')
                name = name_tag.text.strip() if name_tag is not None and name_tag.text else f'{source_filename}_Point'

                description_tag = find_tag(placemark, 'description')
                description = description_tag.text.strip() if description_tag is not None and description_tag.text else ""

                color = "White"
                style_url_tag = find_tag(placemark, 'styleUrl')
                inline_style_tag = find_tag(placemark, 'Style')
                parsed_color_hex = None

                if style_url_tag is not None and style_url_tag.text:
                    style_id = style_url_tag.text.lstrip('#')
                    style_node = root.find(f".//{{*}}Style[@id='{style_id}']") or root.find(
                        f".//{{*}}StyleMap[@id='{style_id}']//{{*}}Style")
                    if style_node is not None:
                        for style_type_node in [find_tag(style_node, 'IconStyle'), find_tag(style_node, 'LineStyle'),
                                                find_tag(style_node, 'PolyStyle')]:
                            if style_type_node is not None:
                                color_node = find_tag(style_type_node, 'color')
                                if color_node is not None and color_node.text:
                                    kml_color_str = color_node.text.strip().lower()
                                    if len(kml_color_str) == 8:
                                        parsed_color_hex = f"#{kml_color_str[6:8]}{kml_color_str[4:6]}{kml_color_str[2:4]}"
                                        break

                elif inline_style_tag is not None:
                    for style_type_node in [find_tag(inline_style_tag, 'IconStyle'),
                                            find_tag(inline_style_tag, 'LineStyle'),
                                            find_tag(inline_style_tag, 'PolyStyle')]:
                        if style_type_node is not None:
                            color_node = find_tag(style_type_node, 'color')
                            if color_node is not None and color_node.text:
                                kml_color_str = color_node.text.strip().lower()
                                if len(kml_color_str) == 8:
                                    parsed_color_hex = f"#{kml_color_str[6:8]}{kml_color_str[4:6]}{kml_color_str[2:4]}"
                                    break

                if parsed_color_hex:
                    color = self.convert_color(parsed_color_hex, "name", True)

                item_data = {"name": name, "color": color, "description": description, "source_file": source_filename}

                point_coords_tag = find_tag(placemark, 'Point/coordinates')
                linestring_coords_tag = find_tag(placemark, 'LineString/coordinates')
                polygon_outer_coords_tag = find_tag(placemark, 'Polygon/outerBoundaryIs/LinearRing/coordinates')

                if point_coords_tag is not None and point_coords_tag.text:
                    parts = point_coords_tag.text.strip().split(',')
                    if len(parts) >= 2:
                        try:
                            lon, lat = float(parts[0]), float(parts[1])
                            alt = float(parts[2]) if len(parts) > 2 else 0.0
                            item_data.update({"lat": lat, "lon": lon, "type": "Landmark", "geometry_type": "Point",
                                              "original_location_data": {"alt": alt}})
                            result.append(item_data)
                        except ValueError:
                            pass

                elif linestring_coords_tag is not None and linestring_coords_tag.text:
                    points_data = [{'lon': float(p.split(',')[0]), 'lat': float(p.split(',')[1]),
                                    'alt': float(p.split(',')[2]) if len(p.split(',')) > 2 else 0.0} for p in
                                   linestring_coords_tag.text.strip().split() if len(p.split(',')) >= 2]
                    if len(points_data) >= 2:
                        item_data.update(
                            {"type": "LineString", "geometry_type": "LineString", "points_data": points_data})
                        result.append(item_data)

                elif polygon_outer_coords_tag is not None and polygon_outer_coords_tag.text:
                    points_data = [{'lon': float(p.split(',')[0]), 'lat': float(p.split(',')[1]),
                                    'alt': float(p.split(',')[2]) if len(p.split(',')) > 2 else 0.0} for p in
                                   polygon_outer_coords_tag.text.strip().split() if len(p.split(',')) >= 2]
                    if len(points_data) >= 3:
                        item_data.update({"type": "Area", "geometry_type": "Polygon", "points_data": points_data})
                        result.append(item_data)

        except ET.ParseError as e:
            self._update_status(f"Помилка парсингу KML: {source_filename} ({e})", error=True)
            return None
        except Exception as e:
            self._update_status(f"Загальна помилка читання KML: {source_filename} ({e})", error=True)
            return None

        return result if result else None

    def read_kml(self, path):
        try:
            with open(path, "r", encoding="utf-8-sig") as fo:
                kml_content = fo.read()
            return self._read_kml_from_content(kml_content, os.path.basename(path))
        except Exception as e:
            self._update_status(f"Помилка KML: {os.path.basename(path)}: {e}", error=True)
            return None

    def read_kmz(self, path):
        try:
            with zipfile.ZipFile(path, 'r') as kmz:
                kml_file_name = next((f for f in kmz.namelist() if f.lower().endswith('.kml')), None)
                if kml_file_name:
                    with kmz.open(kml_file_name) as kml_file_stream:
                        kml_content = kml_file_stream.read().decode('utf-8', errors='replace')
                    return self._read_kml_from_content(kml_content, kml_file_name)
                else:
                    self._update_status(f"У файлі KMZ {os.path.basename(path)} не знайдено .kml файл.", warning=True)
                    return None
        except zipfile.BadZipFile:
            self._update_status(f"Файл {os.path.basename(path)} пошкоджений або не є KMZ архівом.", error=True)
            return None
        except Exception as e:
            self._update_status(f"Помилка KMZ: {os.path.basename(path)}: {e}", error=True)
            return None

    def read_gpx(self, path):
        try:
            namespaces = {
                'gpx': 'http://www.topografix.com/GPX/1/1',
                'gpxx': 'http://www.garmin.com/xmlschemas/GpxExtensions/v3'
            }
            ET.register_namespace('gpxx', namespaces['gpxx'])
            tree = ET.parse(path)
            root = tree.getroot()
            result = []

            for wpt in root.findall('gpx:wpt', namespaces):
                name = wpt.find('gpx:name',
                                namespaces).text if wpt.find('gpx:name', namespaces) is not None else 'GPX Waypoint'
                desc = wpt.find('gpx:desc', namespaces).text if wpt.find('gpx:desc', namespaces) is not None else ''
                color_name = "White"
                try:
                    color_tag = wpt.find('.//gpxx:DisplayColor', namespaces)
                    if color_tag is not None and color_tag.text:
                        color_name = self.convert_color(color_tag.text, "name")
                except Exception:
                    pass

                try:
                    lat, lon = float(wpt.get('lat')), float(wpt.get('lon'))
                    alt = float(wpt.find('gpx:ele', namespaces).text) if wpt.find('gpx:ele',
                                                                                  namespaces) is not None else 0.0
                    result.append({
                        "name": name, "lat": lat, "lon": lon, "type": "Waypoint",
                        "color": color_name, "description": desc,
                        "geometry_type": "Point", "original_location_data": {"alt": alt}
                    })
                except (ValueError, TypeError):
                    continue

            for trk in root.findall('gpx:trk', namespaces):
                trk_name = trk.find('gpx:name',
                                    namespaces).text if trk.find('gpx:name', namespaces) is not None else 'GPX Track'
                for i, trkseg in enumerate(trk.findall('gpx:trkseg', namespaces)):
                    points_data = []
                    for trkpt in trkseg.findall('gpx:trkpt', namespaces):
                        try:
                            lat, lon = float(trkpt.get('lat')), float(trkpt.get('lon'))
                            alt = float(trkpt.find('gpx:ele', namespaces).text) if trkpt.find('gpx:ele',
                                                                                              namespaces) is not None else 0.0
                            points_data.append({'lon': lon, 'lat': lat, 'alt': alt})
                        except (ValueError, TypeError):
                            continue
                    if len(points_data) >= 2:
                        result.append({
                            "name": f"{trk_name} - Сегмент {i + 1}", "type": "TrackSegment",
                            "geometry_type": "LineString", "color": "Red",
                            "description": "", "points_data": points_data
                        })
            return result if result else None
        except Exception as e:
            self._update_status(f"Помилка читання GPX: {os.path.basename(path)}: {e}", error=True)
            return None

    def read_xlsx(self, path):
        try:
            workbook = openpyxl.load_workbook(path, data_only=True)
            result = []
            for sheet in workbook.worksheets:
                if sheet.max_row < 2: continue
                header = [str(cell.value).lower().strip() if cell.value else '' for cell in sheet[1]]
                lat_aliases = ['lat', 'latitude', 'широта', 'y']
                lon_aliases = ['lon', 'long', 'longitude', 'довгота', 'x']
                name_aliases = ['name', 'title', 'назва', 'id']
                color_aliases = ['color', 'колір', 'цвет']
                try:
                    lat_col = next(i for i, h in enumerate(header) if h in lat_aliases)
                    lon_col = next(i for i, h in enumerate(header) if h in lon_aliases)
                    name_col = next((i for i, h in enumerate(header) if h in name_aliases), -1)
                    color_col = next((i for i, h in enumerate(header) if h in color_aliases), -1)
                except StopIteration:
                    self._update_status(f"XLSX: пропуск аркуша '{sheet.title}' (немає колонок lat/lon)",
                                        warning=True)
                    continue
                for r_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
                    try:
                        lat, lon = float(str(row[lat_col]).replace(',', '.')), float(
                            str(row[lon_col]).replace(',', '.'))
                        if not (-90 <= lat <= 90 and -180 <= lon <= 180): continue
                        name = str(
                            row[name_col]) if name_col != -1 and row[name_col] else f'{sheet.title}_Point_{r_idx - 1}'
                        color_value = str(row[color_col]) if color_col != -1 and row[color_col] else "White"
                        color_name = self.convert_color(color_value, "name")
                        desc_parts = [f"{h.capitalize()}: {v}" for h, v in zip(header, row) if
                                      v is not None and header.index(h) not in [lat_col, lon_col, name_col,
                                                                                color_col]]
                        desc = "; ".join(desc_parts)
                        result.append(
                            {"name": name, "lat": lat, "lon": lon, "type": "XLSX Point", "color": color_name,
                             "description": desc, "geometry_type": "Point"})
                    except (ValueError, TypeError, IndexError):
                        continue
            return result if result else None
        except Exception as e:
            self._update_status(f"Помилка XLSX: {os.path.basename(path)}: {e}", error=True)
            return None

    def read_csv(self, path):
        result = []
        try:
            with open(path, mode='r', encoding='utf-8-sig') as infile:
                try:
                    dialect = csv.Sniffer().sniff(infile.read(2048))
                    infile.seek(0)
                except csv.Error:
                    dialect = 'excel'
                    infile.seek(0)
                reader = csv.DictReader(infile, dialect=dialect)
                if not reader.fieldnames: return None
                h_map = {h.lower().strip(): h for h in reader.fieldnames}
                lat_key = next((h_map[alias] for alias in ['lat', 'latitude', 'широта', 'y'] if alias in h_map),
                               None)
                lon_key = next(
                    (h_map[alias] for alias in ['lon', 'long', 'longitude', 'довгота', 'x'] if alias in h_map),
                    None)
                name_key = next((h_map[alias] for alias in ['name', 'title', 'назва', 'id'] if alias in h_map), None)
                desc_key = next((h_map[alias] for alias in ['desc', 'description', 'опис'] if alias in h_map), None)
                color_key = next((h_map[alias] for alias in ['color', 'колір', 'цвет'] if alias in h_map), None)
                if not lat_key or not lon_key:
                    self._update_status("CSV: Не знайдено колонок для широти/довготи.", error=True)
                    return None
                for i, row in enumerate(reader, 2):
                    try:
                        lat, lon = float(str(row[lat_key]).replace(',', '.')), float(
                            str(row[lon_key]).replace(',', '.'))
                        if not (-90 <= lat <= 90 and -180 <= lon <= 180): continue
                        name = row.get(name_key, f'CSV_Point_{i - 1}')
                        color_value = row.get(color_key, "White")
                        color_name = self.convert_color(color_value, "name")
                        other_cols = [k for k in row.keys() if
                                      k not in [lat_key, lon_key, name_key, color_key, desc_key]]
                        desc = row.get(desc_key, "") + "; ".join(
                            [f"{k}: {row[k]}" for k in other_cols if row[k]])
                        result.append(
                            {"name": name, "lat": lat, "lon": lon, "type": "CSV Point", "color": color_name,
                             "description": desc, "geometry_type": "Point"})
                    except (ValueError, TypeError, KeyError):
                        continue
            return result
        except Exception as e:
            self._update_status(f"Помилка CSV: {os.path.basename(path)}: {e}", error=True)
            return None

    def read_scene(self, path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            result = []
            for item in data.get("scene", {}).get("items", []):
                pos = item.get("position", {})
                if "lat" in pos and "lon" in pos:
                    try:
                        result.append({"name": str(item.get("name", "SCENE Point")), "lon": float(pos["lon"]),
                                       "lat": float(pos["lat"]), "type": str(item.get("type", "Landmark")),
                                       "color": self.convert_color(str(item.get("color", "White")), "name",
                                                                   True),
                                       "description": str(item.get("description", "")), "geometry_type": "Point"})
                    except (ValueError, TypeError):
                        continue
            return result if result else None
        except Exception as e:
            self._update_status(f"Помилка .scene: {os.path.basename(path)}: {e}", error=True)
            return None

    def read_geojson(self, path):
        result = []
        try:
            with open(path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            features = data.get("features", []) if data.get(
                "type") == "FeatureCollection" else [data] if data.get("type") == "Feature" else []
            for idx, feature in enumerate(features):
                if not isinstance(feature, dict) or feature.get("type") != "Feature": continue
                geom, props = feature.get("geometry", {}), feature.get("properties", {})
                if not geom or not props: continue
                name = props.get("name", props.get("title", f"GeoFeature_{idx + 1}"))
                color = self.convert_color(str(props.get("color", props.get("stroke", "#ffffff"))), "name", True)
                desc = props.get("description", "")
                item_base = {"name": str(name), "color": color, "description": str(desc),
                             "source_file": os.path.basename(path)}
                geom_type, coords = geom.get("type"), geom.get("coordinates")
                if geom_type == "Point" and coords and len(coords) >= 2:
                    try:
                        item_base.update(
                            {"lat": float(coords[1]), "lon": float(coords[0]), "type": "Landmark",
                             "geometry_type": "Point"})
                        result.append(item_base)
                    except (ValueError, TypeError):
                        pass
                elif geom_type == "LineString" and coords and len(coords) >= 2:
                    points_data = [{'lon': c[0], 'lat': c[1]} for c in coords if len(c) >= 2]
                    if len(points_data) >= 2:
                        item_base.update(
                            {"type": "LineString", "geometry_type": "LineString", "points_data": points_data})
                        result.append(item_base)
                elif geom_type == "Polygon" and coords and len(coords[0]) >= 3:
                    points_data = [{'lon': c[0], 'lat': c[1]} for c in coords[0] if len(c) >= 2]
                    if len(points_data) >= 3:
                        item_base.update({"type": "Area", "geometry_type": "Polygon", "points_data": points_data})
                        result.append(item_base)
        except Exception as e:
            self._update_status(f"Помилка GeoJSON: {os.path.basename(path)}: {e}", error=True)
            return None
        return result if result else None

    def create_scene(self, contents_list, save_path):
        if not contents_list: return False
        items_data = []
        for item in contents_list:
            if item.get('geometry_type') == 'Point':
                items_data.append(
                    {"color": str(item.get("color", "White")), "creationDate": int(time.time() * 1000),
                     "name": str(item.get("name", "N/A")),
                     "position": {"alt": 0.0, "lat": float(item.get("lat", 0.0)),
                                  "lon": float(item.get("lon", 0.0))},
                     "type": str(item.get("type", "Landmark"))})
        scene_obj = {
            "scene": {"items": items_data, "name": os.path.splitext(os.path.basename(save_path))[0]}, "version": 7}
        try:
            with open(save_path, "w", encoding="UTF-8") as f:
                json.dump(scene_obj, f, ensure_ascii=False, separators=(',', ':'));
                return True
        except IOError:
            return False

    def create_kml(self, contents_list, save_path):
        if not contents_list: return False
        try:
            with open(save_path, "w", encoding="UTF-8") as f:
                f.write(self._create_kml_string(contents_list, os.path.splitext(os.path.basename(save_path))[0]));
                return True
        except IOError:
            return False

    def _create_kml_string(self, contents_list, doc_name):
        kml_doc = ET.Element("kml", xmlns="http://www.opengis.net/kml/2.2")
        document = ET.SubElement(kml_doc, "Document")
        ET.SubElement(document, "name").text = doc_name
        style_map = {}
        for item in contents_list:
            color_name = item.get("color", "White")
            if color_name not in style_map:
                color_hex = self.colors.get(color_name, "#ffffff").lstrip('#')
                kml_color = f"ff{color_hex[4:6]}{color_hex[2:4]}{color_hex[0:2]}"
                style_id = f"style_{color_name}"
                style_map[color_name] = style_id
                style = ET.SubElement(document, "Style", id=style_id)
                icon_style = ET.SubElement(style, "IconStyle")
                ET.SubElement(icon_style, "color").text = kml_color
                icon = ET.SubElement(icon_style, "Icon")
                ET.SubElement(icon, "href").text = "http://maps.google.com/mapfiles/kml/pushpin/ylw-pushpin.png"  # Standard icon
                line_style = ET.SubElement(style, "LineStyle")
                ET.SubElement(line_style, "color").text = kml_color
                ET.SubElement(line_style, "width").text = "2"
                poly_style = ET.SubElement(style, "PolyStyle")
                ET.SubElement(poly_style,
                              "color").text = f"7f{color_hex[4:6]}{color_hex[2:4]}{color_hex[0:2]}"
        for item in contents_list:
            placemark = ET.SubElement(document, "Placemark")
            ET.SubElement(placemark, "name").text = xml_escape(item.get("name", "N/A"))
            if item.get("description"): ET.SubElement(placemark, "description").text = xml_escape(
                item.get("description"))
            ET.SubElement(placemark, "styleUrl").text = f"#{style_map.get(item.get('color', 'White'), 'style_White')}"
            geom_type = item.get("geometry_type")
            if geom_type == "Point":
                point = ET.SubElement(placemark, "Point")
                ET.SubElement(point, "coordinates").text = f"{item.get('lon', 0)},{item.get('lat', 0)},0"
            elif geom_type in ["LineString", "Polygon"]:
                coords_data = item.get('points_data', [])
                coords_str = " ".join(f"{p['lon']},{p['lat']},0" for p in coords_data)
                if geom_type == "Polygon" and coords_data and coords_data[0] != coords_data[-1]:
                    coords_str += f" {coords_data[0]['lon']},{coords_data[0]['lat']},0"
                if geom_type == "LineString":
                    geom = ET.SubElement(placemark, "LineString")
                else:  # Polygon
                    geom = ET.SubElement(placemark, "Polygon")
                    outer = ET.SubElement(geom, "outerBoundaryIs")
                    geom = ET.SubElement(outer, "LinearRing")
                ET.SubElement(geom, "coordinates").text = coords_str
        ET.indent(kml_doc, space="  ")
        return '<?xml version="1.0" encoding="UTF-8"?>\n' + ET.tostring(kml_doc, encoding='unicode')

    def create_kmz(self, contents_list, save_path):
        if not contents_list: return False
        kml_content = self._create_kml_string(contents_list, os.path.splitext(os.path.basename(save_path))[0])
        try:
            with zipfile.ZipFile(save_path, 'w', zipfile.ZIP_DEFLATED) as kmz:
                kmz.writestr('doc.kml', kml_content);
                return True
        except Exception:
            return False

    def create_gpx(self, contents_list, save_path):
        if not contents_list: return False
        gpx = ET.Element('gpx', version="1.1", creator="Nexus", xmlns="http://www.topografix.com/GPX/1/1")
        for item in contents_list:
            if item.get('geometry_type') == 'Point':
                wpt = ET.SubElement(gpx, 'wpt', lat=str(item.get("lat")), lon=str(item.get("lon")))
                ET.SubElement(wpt, 'name').text = item.get("name")
            elif item.get('geometry_type') == 'LineString':
                trk = ET.SubElement(gpx, 'trk')
                ET.SubElement(trk, 'name').text = item.get("name")
                trkseg = ET.SubElement(trk, 'trkseg')
                for p in item.get('points_data', []): ET.SubElement(trkseg, 'trkpt', lat=str(p['lat']),
                                                                    lon=str(p['lon']))
        tree = ET.ElementTree(gpx)
        ET.indent(tree, space="  ")
        try:
            tree.write(save_path, encoding='utf-8', xml_declaration=True);
            return True
        except IOError:
            return False

    def create_xlsx(self, contents_list, save_path, split_by_colors=False):
        if not contents_list: return False
        try:
            workbook = xlsxwriter.Workbook(save_path)
        except xlsxwriter.exceptions.FileCreateError:
            self._update_status(f"Помилка XLSX (файл зайнятий?)", error=True);
            return False
        headers = ["NAME", "LAT", "LON", "TYPE", "COLOR", "DESC", "GEOMETRY_TYPE", "WKT"]
        header_format = workbook.add_format(
            {'bold': True, 'bg_color': '#D9EAD3', 'border': 1, 'align': 'center'})

        def write_sheet(ws, data):
            ws.write_row(0, 0, headers, header_format)
            for r, item in enumerate(data, 1):
                geom_type = item.get("geometry_type")
                wkt = ""
                if geom_type == "Point":
                    wkt = f"POINT ({item.get('lon')} {item.get('lat')})"
                elif geom_type in ["LineString", "Polygon"]:
                    pts = ", ".join(f"{p['lon']} {p['lat']}" for p in item.get('points_data', []))
                    if geom_type == "Polygon" and item.get('points_data'):
                        p_data = item['points_data']
                        if p_data[0]['lon'] != p_data[-1]['lon'] or p_data[0]['lat'] != p_data[-1]['lat']:
                            pts += f", {p_data[0]['lon']} {p_data[0]['lat']}"
                        wkt = f"POLYGON (({pts}))"
                    else:
                        wkt = f"LINESTRING ({pts})"
                row_data = [item.get(k, '') for k in
                            ['name', 'lat', 'lon', 'type', 'color', 'description']] + [geom_type, wkt]
                ws.write_row(r, 0, row_data)
            ws.autofit()

        if split_by_colors:
            data_by_color = {}
            for item in contents_list: data_by_color.setdefault(item.get('color', 'NoColor'), []).append(item)
            for color, data in data_by_color.items(): write_sheet(workbook.add_worksheet(color[:31]), data)
        else:
            write_sheet(workbook.add_worksheet("Data"), contents_list)
        try:
            workbook.close();
            return True
        except xlsxwriter.exceptions.FileCreateError:
            return False
            
    # --- COLOR METHODS ---
    @staticmethod
    def _color_distance(rgb1: Tuple[int, int, int], rgb2: Tuple[int, int, int]) -> float:
        """Calculate Euclidean distance between two RGB colors."""
        return math.sqrt(sum((c1 - c2) ** 2 for c1, c2 in zip(rgb1, rgb2)))

    def _find_closest_color_name(self, rgb_tuple: Tuple[int, int, int]) -> str:
        """Finds the closest color name from the palette."""
        if not isinstance(rgb_tuple, (list, tuple)) or len(rgb_tuple) < 3:
            return "White"
        rgb_tuple = tuple(max(0, min(255, int(c))) for c in rgb_tuple[:3])
        min_dist = float('inf')
        closest_name = "White"
        for name, palette_rgb in self._palette_rgb.items():
            dist = self._color_distance(rgb_tuple, palette_rgb)
            if dist < min_dist:
                min_dist = dist
                closest_name = name
        return closest_name

    def convert_color(self, color_value: Any, target_format: str = 'name', allow_name_lookup_from_hex=False) -> str:
        """
        Приводить колір до різних форматів:
        - 'name': англійська назва кольору (Red, Blue...)
        - 'hex': HEX-рядок (#RRGGBB)
        - 'str_rgb': рядок формату R,G,B
        """
        if not color_value:
            return "White" if target_format == 'name' else self.colors["White"]

        # Directly handle if it's already a valid color name
        if isinstance(color_value, str) and color_value.capitalize() in self.colors:
            color_name_en = color_value.capitalize()
        else:
            rgb_tuple = None
            color_name_en = None
            if isinstance(color_value, (list, tuple)):
                rgb_tuple = color_value
            elif isinstance(color_value, str):
                value_lower = color_value.lower().strip()
                # HEX
                hex_match = re.fullmatch(r'#?([0-9a-fA-F]{6}|[0-9a-fA-F]{3})', value_lower)
                if hex_match:
                    hex_str = hex_match.group(1)
                    if len(hex_str) == 3:
                        hex_str = "".join([c * 2 for c in hex_str])
                    rgb_tuple = (int(hex_str[0:2], 16), int(hex_str[2:4], 16), int(hex_str[4:6], 16))
                else:
                    # RGBA
                    rgba_match = re.match(r'rgba?\((\d{1,3}),\s*(\d{1,3}),\s*(\d{1,3})', value_lower)
                    if rgba_match:
                        rgb_tuple = (int(rgba_match.group(1)), int(rgba_match.group(2)), int(rgba_match.group(3)))
                    else:
                        for keyword, en_name in self.color_keyword_map.items():
                            if keyword in value_lower:
                                color_name_en = en_name
                                break
            if rgb_tuple:
                color_name_en = self._find_closest_color_name(rgb_tuple)
            if not color_name_en:
                print(f"[DEBUG] Не розпізнано колір: {color_value}, повертаю White")
                color_name_en = "White"

        # Return requested format
        if target_format == 'name':
            return color_name_en
        elif target_format == 'hex':
            return self.colors.get(color_name_en, self.colors["White"])
        elif target_format == 'str_rgb':
            h = self.colors.get(color_name_en, self.colors["White"]).lstrip('#')
            return f"{int(h[0:2], 16)},{int(h[2:4], 16)},{int(h[4:6], 16)}"
        return color_name_en

    def color_to_csv_rgba(self, color_value):
        """
        Повертає рядок формату rgba(R,G,B,1) для заданого кольору.
        """
        hex_color = self.convert_color(color_value, target_format='hex')
        if hex_color.startswith('#'):
            hex_color = hex_color[1:]
        r = int(hex_color[0:2], 16)
        g = int(hex_color[2:4], 16)
        b = int(hex_color[4:6], 16)
        return f"rgba({r},{g},{b},1)"
        
    # --- NEW UNIFIED CSV LOGIC ---
    def create_csv(self, contents_list: List[Dict[str, Any]], base_save_path: str) -> bool:
            """
            Створює єдиний CSV файл для всіх типів геометрії (точки, лінії, полігони),
            зберігаючи SIDC та інформацію про стиль у форматі, сумісному з еталоном.
            """
            if not contents_list:
                self._update_status("Немає даних для запису в CSV.", warning=True)
                return False

            self._update_status(f"Створення універсального CSV: {os.path.basename(base_save_path)}...",
                                self.C_BUTTON_HOVER)

            headers = ["sidc", "id", "quantity", "name", "observation_datetime", "reliability_credibility",
                       "staff_comments", "platform_type", "direction", "speed", "coordinates", "comment 1", "comment 2",
                       "comment 3", "comment 4"]

            try:
                for chunk_index, i in enumerate(range(0, len(contents_list), self.CSV_CHUNK_SIZE)):
                    chunk_contents = contents_list[i:i + self.CSV_CHUNK_SIZE]
                    current_save_path = self._get_chunked_save_path(base_save_path, chunk_index)

                    with open(current_save_path, "w", newline='', encoding="utf-8") as csvfile:
                        writer = csv.writer(csvfile)
                        writer.writerow(headers)
                        for item in chunk_contents:
                            row = [""] * len(headers)
                            name = item.get('name', '')
                            observation_datetime = ''
                            geom_type = item.get('geometry_type')

                            row[3] = name
                            row[4] = observation_datetime

                            # --- Логіка для Точок ---
                            if geom_type == 'Point':
                                wkt = f"POINT ({item.get('lon', 0.0)} {item.get('lat', 0.0)})"
                                color_hex = self.convert_color(item.get("color"), 'hex')

                                row[0] = item.get('milgeo:meta:sidc') or POINT_SIDC
                                row[10] = wkt
                                row[11] = color_hex  # comment 1 для кольору у форматі HEX

                            # --- Логіка для Ліній та Полігонів ---
                            elif geom_type in ['LineString', 'Polygon']:
                                points_data = item.get('points_data', [])
                                if not points_data:
                                    continue

                                coords_parts = [f"{p.get('lon', 0.0)} {p.get('lat', 0.0)}" for p in points_data]

                                if geom_type == 'Polygon':
                                    if coords_parts and coords_parts[0] != coords_parts[-1]:
                                        coords_parts.append(coords_parts[0])
                                    wkt = f"POLYGON (({', '.join(coords_parts)}))"
                                else:  # LineString
                                    wkt = f"LINESTRING ({', '.join(coords_parts)})"

                                # Визначення кольору та SIDC
                                color_hex = self.convert_color(item.get('color'), 'hex')
                                sidc = get_line_sidc(color_hex)
                                row[0] = item.get('milgeo:meta:sidc') or sidc
                                row[10] = wkt
                                # Додавання стилів у коментарі
                                row[11] = "stroke-opacity: 1"
                                row[12] = f"stroke: {color_hex}"
                                row[13] = "stroke-width: 3"
                                row[14] = "icon-scale: 0"
                            else:
                                continue  # Пропустити невідомі типи геометрії

                            writer.writerow(row)

                self._update_status(f"Файл CSV успішно збережено: {os.path.basename(base_save_path)}",
                                    self.C_ACCENT_DONE)
                return True
            except Exception as e:
                self._update_status(f"Помилка під час створення CSV: {e}", error=True)
                import traceback
                traceback.print_exc()
                return False

    def create_geojson(self, contents_list, save_path):
        if not contents_list: return False
        features = []
        for item in contents_list:
            geom_type = item.get('geometry_type')
            props = {k.replace("milgeo:meta:", ""): v for k, v in item.items() if
                     k.startswith("milgeo:meta:") and v is not None}
            props.update({k: v for k, v in item.items() if
                          not k.startswith("milgeo:meta:") and k not in ['points_data', 'geometry_type',
                                                                         'original_location_data'] and v is not None})
            props['color'] = self.convert_color(item.get("color", "White"), "hex", True)
            geometry = None
            if geom_type == 'Point':
                geometry = {"type": "Point", "coordinates": [item.get("lon", 0.0), item.get("lat", 0.0),
                                                             item.get("original_location_data", {}).get("alt",
                                                                                                        0.0) or 0.0]}
            elif geom_type in ['LineString', 'Polygon']:
                coords = [[p.get('lon', 0.0), p.get('lat', 0.0)] for p in item.get('points_data', [])]
                if len(coords) >= (2 if geom_type == 'LineString' else 3):
                    if geom_type == 'Polygon' and coords[0] != coords[-1]: coords.append(coords[0])
                    geometry = {"type": geom_type, "coordinates": [coords] if geom_type == 'Polygon' else coords}
            if geometry: features.append({"type": "Feature", "properties": props, "geometry": geometry})
        if not features: return False
        try:
            with open(save_path, "w", encoding="UTF-8") as f:
                json.dump({"type": "FeatureCollection", "features": features}, f, indent=2,
                          ensure_ascii=False);
                return True
        except IOError:
            return False
    
    # --- МЕТОДИ ДЛЯ РОБОТИ З КОЛЬОРОМ ---
    @staticmethod
    def _color_distance(rgb1: Tuple[int, int, int], rgb2: Tuple[int, int, int]) -> float:
        """Обчислює евклідову відстань між двома кольорами RGB."""
        return math.sqrt(sum([(c1 - c2) ** 2 for c1, c2 in zip(rgb1, rgb2)]))

    def _find_closest_color_name(self, rgb_tuple: Tuple[int, int, int]) -> str:
        """Знаходить найближчу назву кольору з палітри."""
        if not isinstance(rgb_tuple, (list, tuple)) or len(rgb_tuple) < 3:
            return "White"
        rgb_tuple = tuple(max(0, min(255, c)) for c in rgb_tuple[:3])
        min_dist = float('inf')
        closest_name = "White"
        for name, palette_rgb in self._palette_rgb.items():
            dist = self._color_distance(rgb_tuple, palette_rgb)
            if dist < min_dist:
                min_dist = dist
                closest_name = name
        return closest_name

    def convert_color(self, color_value: Any, target_format: str = 'name', allow_name_lookup_from_hex=False) -> str:
        """Надійно конвертує представлення кольору (назва, hex, rgb) у стандартизований колір з палітри."""
        if not color_value:
            return "White" if target_format == 'name' else self.colors["White"]

        if isinstance(color_value, str) and color_value.capitalize() in self.colors:
            color_name_en = color_value.capitalize()
        else:
            rgb_tuple = None
            color_name_en = None
            if isinstance(color_value, (list, tuple)):
                rgb_tuple = color_value
            elif isinstance(color_value, str):
                value_lower = color_value.lower().strip()
                hex_match = re.search(r'#?([0-9a-fA-F]{6}|[0-9a-fA-F]{3})\b', value_lower)
                if hex_match:
                    hex_str = hex_match.group(1)
                    if len(hex_str) == 3:
                        hex_str = "".join([c * 2 for c in hex_str])
                    rgb_tuple = (int(hex_str[0:2], 16), int(hex_str[2:4], 16), int(hex_str[4:6], 16))
                else:
                    rgba_match = re.search(r'rgba?\((\d{1,3}),\s*(\d{1,3}),\s*(\d{1,3})', value_lower)
                    if rgba_match:
                        rgb_tuple = (int(rgba_match.group(1)), int(rgba_match.group(2)), int(rgba_match.group(3)))
                    else:
                        for keyword, en_name in self.color_keyword_map.items():
                            if keyword in value_lower:
                                color_name_en = en_name
                                break
            if rgb_tuple:
                color_name_en = self._find_closest_color_name(rgb_tuple)
            if not color_name_en:
                color_name_en = "White"

        if target_format == 'name':
            return color_name_en
        elif target_format == 'hex':
            return self.colors.get(color_name_en, self.colors["White"])
        elif target_format == 'str_rgb':
            h = self.colors.get(color_name_en, self.colors["White"]).lstrip('#')
            return f"{int(h[0:2], 16)},{int(h[2:4], 16)},{int(h[4:6], 16)}"
        return color_name_en

    def _apply_selected_numeration(self, point_list):
        if not point_list: return []
        if len(point_list) == 1:
            point_list[0]["name"] = self.generate_free_numbers_list(1)[0];
            return point_list
        method = self.numerations.index(self.chosen_numeration.get())
        if method == 0: return self.apply_neighbor_numeration(point_list)
        if method == 1: return self.apply_snake_numeration(point_list)
        if method == 2: return self.apply_two_axis_numeration(point_list)
        if method == 3: return self.apply_one_axis_numeration(point_list)
        if method == 4: return self.apply_random_numeration(point_list)
        return point_list

    def generate_free_numbers_list(self, count):
        if count <= 0: return []
        numbers, num, exceptions_on = [], 1, self.exceptions_agree.get()
        while len(numbers) < count:
            if not (exceptions_on and (30 <= num <= 40 or 500 <= num <= 510)):
                numbers.append(str(num))
            num += 1
        return numbers

    def apply_random_numeration(self, content_list):
        if not content_list: return []
        result = copy.deepcopy(content_list)
        free_numbers = self.generate_free_numbers_list(len(result))
        random.shuffle(free_numbers)
        for i, item in enumerate(result): item["name"] = free_numbers[i]
        return result

    def calculate_distance(self, p1, p2):
        R = 6371e3
        lat1, lon1, lat2, lon2 = map(math.radians, [p1[0], p1[1], p2[0], p2[1]])
        dlon, dlat = lon2 - lon1, lat2 - lat1
        a = sin(dlat / 2) ** 2 + cos(lat1) * cos(lat2) * sin(dlon / 2) ** 2
        return R * 2 * atan2(sqrt(a), sqrt(1 - a))

    def apply_snake_numeration(self, content_list: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Присвоює порядкові номери точкам у стилі "змійка": рядок за рядком."""
        if not content_list:
            return content_list
        points = [p for p in content_list if 'lat' in p and 'lon' in p]
        if not points:
            return content_list
        min_lon, max_lon = min(p['lon'] for p in points), max(p['lon'] for p in points)
        min_lat, max_lat = min(p['lat'] for p in points), max(p['lat'] for p in points)
        lat_range = max_lat - min_lat if max_lat != min_lat else 1e-6  # Avoid division by zero
        points.sort(key=lambda p: (int((p['lat'] - min_lat) / lat_range * 10),
                                   p['lon'] if int((p['lat'] - min_lat) / lat_range * 10) % 2 == 0 else -p['lon']))
        free_numbers = self.generate_free_numbers_list(len(points))
        for i, item in enumerate(points):
            item['name'] = free_numbers[i]
        return points

    def apply_neighbor_numeration(self, content_list):
        if not content_list: return []
        points = [p for p in content_list if 'lat' in p and 'lon' in p]
        if not points: return content_list
        unvisited = points[:]
        start_point = min(unvisited, key=lambda p: (p['lat'], p['lon']))
        ordered_points = [start_point]
        unvisited.remove(start_point)
        current_point = start_point
        while unvisited:
            next_point = min(unvisited,
                             key=lambda p: self.calculate_distance((current_point['lat'], current_point['lon']),
                                                                   (p['lat'], p['lon'])))
            ordered_points.append(next_point)
            unvisited.remove(next_point)
            current_point = next_point
        free_numbers = self.generate_free_numbers_list(len(ordered_points))
        for i, item in enumerate(ordered_points):
            item['name'] = free_numbers[i]
        return ordered_points

    def apply_one_axis_numeration(self, content_list):
        if not content_list: return []
        points = [p for p in content_list if 'lat' in p and 'lon' in p]
        if not points: return content_list
        trans = self.chosen_translation.get()
        if trans == "На 90 градусів" or trans == "На 270 градусів":
            key, reverse = 'lat', trans == "На 270 градусів"
        else:
            key, reverse = 'lon', trans == "На 180 градусів"
        points.sort(key=lambda p: p[key], reverse=reverse)
        free_numbers = self.generate_free_numbers_list(len(points))
        for i, item in enumerate(points): item['name'] = free_numbers[i]
        return points

    def apply_two_axis_numeration(self, content_list):
        if not content_list: return []
        points = [p for p in content_list if 'lat' in p and 'lon' in p]
        if not points: return content_list
        min_lat, max_lat = min(p['lat'] for p in points), max(p['lat'] for p in points)
        min_lon, max_lon = min(p['lon'] for p in points), max(p['lon'] for p in points)
        trans = self.chosen_translation.get()
        if trans == "Не повертати":
            corner = (min_lat, min_lon)
        elif trans == "На 90 градусів":
            corner = (min_lat, max_lon)
        elif trans == "На 180 градусів":
            corner = (max_lat, max_lon)
        else:
            corner = (max_lat, min_lon)
        points.sort(key=lambda p: self.calculate_distance(corner, (p['lat'], p['lon'])))
        free_numbers = self.generate_free_numbers_list(len(points))
        for i, item in enumerate(points): item['name'] = free_numbers[i]
        return points


if __name__ == "__main__":
    try:
        # Для кращого відображення на Windows
        ctypes.windll.shcore.SetProcessDpiAwareness(1)
    except (ImportError, AttributeError, OSError):
        pass  # Не спрацює на інших ОС, і це нормально
    app = Main()
    app.run()
